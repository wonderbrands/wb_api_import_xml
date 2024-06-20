from email.message import EmailMessage
from email.utils import make_msgid
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from pprint import pprint
from email import encoders
from tqdm import tqdm
import time
import json
# import jsonrpc
# import jsonrpclib
# import random
# import urllib.request
# import getpass
# import http
# import requests
# import logging
# import zipfile
# import socket
import os
import locale
import xmlrpc.client
import base64
import openpyxl
# import xlrd
# import pandas as pd
# import MySQLdb
import mysql.connector
import smtplib
# import ssl
# import email
import datetime

from Test import extract_orders as e_o

print('================================================================')
print('BIENVENIDO AL PROCESO DE NOTAS DE CRÉDITO PARA MARKETPLACES')
print('================================================================')
print('SCRIPT DE CREACIÓN DE NOTAS DE CRÉDITO')
print('================================================================')
today_date = datetime.datetime.now()
dir_path = os.path.dirname(os.path.realpath(__file__))
print('Fecha:' + today_date.strftime("%Y-%m-%d %H:%M:%S"))
#Archivo de configuración - Use config_dev.json si está haciendo pruebas
#Archivo de configuración - Use config.json cuando los cambios vayan a producción

# ***********************************************
# ARCHIVO DE CONFIGURACIÓN
config_file = 'config.json'
# ***********************************************

config_file_name = rf'C:\Users\Sergio Gil Guerrero\Documents\WonderBrands\Repos\wb_odoo_external_api\config\{config_file}'
l10n_mx_edi_payment_method_id = 3
l10n_mx_edi_usage = 'G02'

#FECHAS DEL PERIODO ***********************************************
start_date_str = datetime.date(2024, 1, 1).strftime("%Y-%m-%d")
end_date_str = datetime.date(2024, 6, 1).strftime("%Y-%m-%d")
month_executed = 'Mayo'
# ***********************************************

#PATHS de los archivos de ordenes conciliadas
orders_meli_file_path = 'C:/Users/Sergio Gil Guerrero/Documents/WonderBrands/Finanzas/{}/Conciliadas/Notas_de_credito_totales_ML.csv'.format(month_executed)
#orders_meli_file_path = 'C:/Users/Sergio Gil Guerrero/Documents/WonderBrands/Finanzas/{}/Conciliadas/NC-pendientes.csv'.format(month_executed)
orders_amz_file_path = 'C:/Users/Sergio Gil Guerrero/Documents/WonderBrands/Finanzas/{}/Conciliadas/Notas_de_credito_totales_AMZ.csv'.format(month_executed)

def get_odoo_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['odoo']
def get_psql_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['psql']
def get_email_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['email']
def reverse_invoice_meli(): #NOTAS DE CRÉDITO INDIVIDUALES MELI
    #Formato para query
    type_filter = 'INDIVIDUAL'
    marketplace_filter = 'MERCADO LIBRE'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_meli_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str, start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO INDIVIDUALES MELI')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    #INDIVIDUALES MELI
    mycursor.execute("""#INDIVIDUALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               ifnull(d.payment_date_last_modified, dd.payment_date_last_modified) 'payment_date_last_modified'/*,
                               ifnull(d.order_id, dd.pack_id) 'order_id_or_pack_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               ifnull(d.refunded_amt, dd.refunded_amt) 'ml_refunded_amount',
                               b.invoice_partner_display_name 'cliente',
                               b.name,
                               'INDIVIDUAL' as type,
                               'MERCADO LIBRE' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON b.invoice_origin = c.name
                        LEFT JOIN (SELECT a.order_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(paid_amt) 'paid_amt', SUM(refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                                   FROM somos_reyes.ml_order_payments a
                                   LEFT JOIN somos_reyes.ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT a.pack_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(b.paid_amt) 'paid_amt', SUM(b.refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                        FROM somos_reyes.ml_order_update a
                        LEFT JOIN somos_reyes.ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                        GROUP BY 1) dd
                        ON c.yuju_pack_id = dd.pack_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE (d.order_id is not null or dd.pack_id is not null)
                        AND e.invoice_origin is null
                        AND ((ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total < 1 AND ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total > -1)
                        OR (ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total < 1
                        AND ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total > -1))
                        AND c.name in ({});
                            """.format(placeholders), tuple(dates_list_params+list_orders))
    invoice_records = mycursor.fetchall()
    # Lista de SO a las que se les creó una credit_notes
    so_modified = []
    # Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    # Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    # Lista de nombres de las notas de crédito creadas
    nc_created = []
    # Lista de facturas origen
    so_origin_invoice = []
    # Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    # Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0] #Es la SO
            inv_id = each[1]
            nc_date = each[2].strftime("%Y-%m-%d %H:%M:%S")
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['invoice_origin', '=', inv_origin_name]]])
            if invoice:
                # Se verifica si ya existe una nota de crédito para esta orden de venta
                existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                if not existing_credit_note:
                #if 'out_refund' not in inv_move_types:
                    for inv in invoice:
                        inv_id = inv['id'] # ID de la factura
                        inv_name = inv['name'] # Nombre de la factura
                        inv_origin = inv['invoice_origin'] # Nombre de la SO ligada a la factura
                        #inv_narration = inv['narration']
                        #inv_uuid = inv_narration[3:-4]
                        inv_uuid = inv['l10n_mx_edi_cfdi_uuid'] # Folio fiscal de la factura
                        inv_journal_id = inv['journal_id'][0] #Diario de la factura
                        inv_state = inv['state']
                        l10n_mx_edi_origin = '03|' + str(inv_uuid)
                        team_id = inv['team_id'][0]

                        if inv_state == 'posted':
                            #Se hace una llamada al wizard de creación de notas de crédito
                            credit_note_wizard = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'create',
                                                                   [{
                                'refund_method': 'refund',
                                'reason': 'Por efectos de devolución o retorno de una orden',
                                'journal_id': inv_journal_id, }],
                                           {'context': {
                                               'active_ids': [inv_id],
                                               'active_id': inv_id,
                                               'active_model': 'account.move',
                                           }}
                                        )
                            #Se crea la nota de crédito con la info anterior y se usa la función reverse_moves del botón revertir en el wizard
                            nc_inv_create = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'reverse_moves',[credit_note_wizard])
                            nc_id = nc_inv_create['res_id'] # Obtiene el id de la nota de crédito
                            # Agrega un mensaje al chatter de la nota de crédito
                            message = {
                                'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {inv_origin}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                'message_type': 'comment',
                            }
                            write_msg_tech = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[nc_id], message)
                            #Actualización de Forma de Pago, CFDI Origen, Equipo de Ventas
                            update_vals_nc = models.execute_kw(db_name, uid, password, 'account.move', 'write',[[nc_id], {'team_id': team_id, 'l10n_mx_edi_origin': l10n_mx_edi_origin, 'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id, 'l10n_mx_edi_usage': l10n_mx_edi_usage}])
                            #Confirma la nota de crédito
                            #upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post',[nc_id])
                            # Timbramos la nota de crédito
                            # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[nc_id])
                            #buscamos el nombre de la nota creada
                            search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', nc_id]]])
                            nc_name = search_nc_name[0]['name']
                            nc_total = search_nc_name[0]['amount_total']
                            sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read',[[['name', '=', inv_origin_name]]])[0]
                            sale_ref = sale_order['channel_order_reference']
                            #Agregamos a las listas
                            nc_created.append(nc_name)
                            nc_amount_total.append(nc_total)
                            so_modified.append(inv_origin)
                            so_origin_invoice.append(inv_name)
                            so_mkp_reference.append(sale_ref)
                            progress_bar.update(1)
                        else:
                            print(f"La factura de la órden {inv_origin_name} no está confirmada")
                            progress_bar.update(1)
                            continue
                else:
                    print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                    so_with_refund.append(inv_origin_name)
                    progress_bar.update(1)
                    continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_individuales_meli_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales 
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito 
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron 
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['sergio@wonderbrands.co', 'eric@wonderbrands.co','rosalba@wonderbrands.co', 'natalia@wonderbrands.co', 'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Meli - Creación de notas de crédito para facturas individuales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('PROCESO NC PARA MERCADO LIBRE COMPLETADO :)')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
def reverse_invoice_global_meli(): #NOTAS DE CRÉDITO GLOBALES MELI
    # Formato para query
    type_filter = 'GLOBAL'
    marketplace_filter = 'MERCADO LIBRE'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_meli_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str, start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO GLOBALES MELI')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    # GLOBALES MELI
    mycursor.execute("""#GLOBALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name/*,
                               ifnull(d.order_id, dd.pack_id) 'order_id_or_pack_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               ifnull(d.refunded_amt, dd.refunded_amt) 'ml_refunded_amount',
                               ifnull(d.payment_date_last_modified, dd.payment_date_last_modified) 'payment_date_last_modified',
                               b.invoice_partner_display_name 'cliente',
                               'GLOBAL' as type,
                               'MERCADO LIBRE' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON SUBSTRING_INDEX(SUBSTRING_INDEX(invoice_ids, ']', 1), '[', -1) = b.id
                        LEFT JOIN (SELECT a.order_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(paid_amt) 'paid_amt', SUM(refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                                   FROM somos_reyes.ml_order_payments a
                                   LEFT JOIN somos_reyes.ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT a.pack_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(b.paid_amt) 'paid_amt', SUM(b.refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                        FROM somos_reyes.ml_order_update a
                        LEFT JOIN somos_reyes.ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                        GROUP BY 1) dd
                        ON c.yuju_pack_id = dd.pack_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE (d.order_id is not null or dd.pack_id is not null)
                        AND e.invoice_origin is null
                        AND invoice_partner_display_name = 'PÚBLICO EN GENERAL'
                        AND (ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total > 1 OR ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total < -1)
                        AND (ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total > 1 OR ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total < -1)
                        AND ((ifnull(d.refunded_amt, dd.refunded_amt) - c.amount_total < 1 AND ifnull(d.refunded_amt, dd.refunded_amt) - c.amount_total > -1)
                        OR (ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - c.amount_total < 1
                        AND ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - c.amount_total > -1))
                        AND c.name in ({});""".format(placeholders), tuple(dates_list_params+list_orders))
    invoice_records = mycursor.fetchall()
    #invoice_records = [('SO2789803','1548302','INV/8202/65043')]
    # invoice_records = [('SO2669797', '1457583', 'INV/8202/59620'), ('SO2779433', '1541845', 'INV/8202/63880'), ('SO2779433', '1541845', 'INV/8202/63880'), ('SO2789803', '1548302', 'INV/8202/65043'), ('SO2860100', '1653481', 'INV/8202/70549'), ('SO2852951', '1638257', 'INV/8202/68345'), ('SO2852943', '1638257', 'INV/8202/68345'), ('SO2845876', '1629989', 'INV/8202/67510'), ('SO2843696', '1630529', 'INV/8202/67516'), ('SO2841212', '1617569', 'INV/8202/67205'), ('SO2834870', '1650259', 'INV/8202/70374'), ('SO2832919', '1617769', 'INV/8202/67211'), ('SO2829212', '1649886', 'INV/8202/69897'), ('SO2829212', '1649886', 'INV/8202/69897'), ('SO2822418', '1611246', 'INV/8202/67283'), ('SO2817695', '1611240', 'INV/8202/66920'), ('SO2817694', '1606752', 'INV/8202/66812'), ('SO2817683', '1606752', 'INV/8202/66812'), ('SO2812286', '1606807', 'INV/8202/66827'), ('SO2811523', '1606807', 'INV/8202/66827'), ('SO2808987', '1606739', 'INV/8202/66807'), ('SO2932827', '1737657', 'INV/8202/75258'), ('SO2932110', '1731666', 'INV/8202/75127'), ('SO2931657', '1731682', 'INV/8202/75131'), ('SO2931549', '1731682', 'INV/8202/75131'), ('SO2930612', '1731682', 'INV/8202/75131'), ('SO2930236', '1731682', 'INV/8202/75131'), ('SO2929655', '1731682', 'INV/8202/75131'), ('SO2928279', '1731680', 'INV/8202/75130'), ('SO2928137', '1728198', 'INV/8202/74899'), ('SO2926202', '1731680', 'INV/8202/75130'), ('SO2924578', '1728234', 'INV/8202/74903'), ('SO2923970', '1731961', 'INV/8202/75208'), ('SO2923970', '1731961', 'INV/8202/75208'), ('SO2922949', '1731961', 'INV/8202/75208'), ('SO2922949', '1731961', 'INV/8202/75208'), ('SO2922095', '1731680', 'INV/8202/75130'), ('SO2921244', '1727704', 'INV/8202/74759'), ('SO2921043', '1729188', 'INV/8202/74919'), ('SO2920928', '1731680', 'INV/8202/75130'), ('SO2920928', '1731680', 'INV/8202/75130'), ('SO2919597', '1727702', 'INV/8202/74761'), ('SO2919495', '1727702', 'INV/8202/74761'), ('SO2919153', '1716309', 'INV/8202/74496'), ('SO2918587', '1716303', 'INV/8202/74494'), ('SO2918429', '1716309', 'INV/8202/74496'), ('SO2918055', '1716309', 'INV/8202/74496'), ('SO2917949', '1716311', 'INV/8202/74497'), ('SO2917943', '1716311', 'INV/8202/74497'), ('SO2915370', '1716275', 'INV/8202/74490'), ('SO2915369', '1716275', 'INV/8202/74490'), ('SO2915339', '1716275', 'INV/8202/74490'), ('SO2914980', '1731961', 'INV/8202/75208'), ('SO2914980', '1731961', 'INV/8202/75208'), ('SO2914892', '1724096', 'INV/8202/74697'), ('SO2914768', '1724096', 'INV/8202/74697'), ('SO2913519', '1716275', 'INV/8202/74490'), ('SO2912786', '1724134', 'INV/8202/74699'), ('SO2912786', '1724134', 'INV/8202/74699'), ('SO2912144', '1724134', 'INV/8202/74699'), ('SO2911826', '1716288', 'INV/8202/74492'), ('SO2911024', '1716290', 'INV/8202/74620'), ('SO2909512', '1716302', 'INV/8202/74495'), ('SO2907689', '1712892', 'INV/8202/74008'), ('SO2906148', '1716291', 'INV/8202/74493'), ('SO2906005', '1724134', 'INV/8202/74699'), ('SO2906003', '1724134', 'INV/8202/74699'), ('SO2903841', '1710999', 'INV/8202/73963'), ('SO2902232', '1716291', 'INV/8202/74493'), ('SO2902220', '1731683', 'INV/8202/75133'), ('SO2899282', '1712903', 'INV/8202/74012'), ('SO2899227', '1712909', 'INV/8202/74016'), ('SO2898904', '1712903', 'INV/8202/74012'), ('SO2898560', '1724164', 'INV/8202/74701'), ('SO2896784', '1708972', 'INV/8202/73537'), ('SO2893408', '1710835', 'INV/8202/73934'), ('SO2890887', '1710857', 'INV/8202/74000'), ('SO2890242', '1713733', 'INV/8202/74087'), ('SO2886042', '1710872', 'INV/8202/73943'), ('SO2881006', '1710892', 'INV/8202/73947'), ('SO2880849', '1708692', 'INV/8202/73397'), ('SO2872690', '1716436', 'INV/8202/74509'), ('SO2870928', '1708830', 'INV/8202/73422'), ('SO2867925', '1713840', 'INV/8202/74151'), ('SO2867386', '1708895', 'INV/8202/73432'), ('SO2865400', '1667712', 'INV/8202/71620'), ('SO2858934', '1708965', 'INV/8202/73440'), ('SO3017108', '1846874', 'INV/8202/79185'), ('SO3020583', '1840853', 'INV/8202/78567'), ('SO3019799', '1840920', 'INV/8202/78595'), ('SO3019793', '1847009', 'INV/8202/79196'), ('SO2976401', '1840853', 'INV/8202/78567'), ('SO3003687', '1847017', 'INV/8202/79202'), ('SO3018102', '1847017', 'INV/8202/79202'), ('SO3017690', '1847017', 'INV/8202/79202'), ('SO3017667', '1840853', 'INV/8202/78567'), ('SO3017472', '1840877', 'INV/8202/78577'), ('SO3017304', '1840930', 'INV/8202/78599'), ('SO3017151', '1840877', 'INV/8202/78577'), ('SO3016868', '1846855', 'INV/8202/79184'), ('SO3016791', '1840877', 'INV/8202/78577'), ('SO3016633', '1840877', 'INV/8202/78577'), ('SO3016516', '1840920', 'INV/8202/78595'), ('SO3015943', '1840901', 'INV/8202/78585'), ('SO3015784', '1840901', 'INV/8202/78585'), ('SO3015659', '1840901', 'INV/8202/78585'), ('SO3015358', '1840829', 'INV/8202/78552'), ('SO3014544', '1840901', 'INV/8202/78585'), ('SO3015266', '1847017', 'INV/8202/79202'), ('SO3014777', '1840901', 'INV/8202/78585'), ('SO3013333', '1830420', 'INV/8202/77776'), ('SO3014526', '1840914', 'INV/8202/78591'), ('SO3014442', '1840914', 'INV/8202/78591'), ('SO3012981', '1840829', 'INV/8202/78552'), ('SO3012919', '1840829', 'INV/8202/78552'), ('SO3012680', '1840930', 'INV/8202/78599'), ('SO3001437', '1840935', 'INV/8202/78601'), ('SO3011914', '1840935', 'INV/8202/78601'), ('SO3011228', '1840935', 'INV/8202/78601'), ('SO3010244', '1826331', 'INV/8202/77083'), ('SO3009240', '1840947', 'INV/8202/78607'), ('SO3009217', '1840852', 'INV/8202/78565'), ('SO3009212', '1847587', 'INV/8202/79294'), ('SO3009208', '1847587', 'INV/8202/79294'), ('SO3009168', '1826315', 'INV/8202/77078'), ('SO3009123', '1840947', 'INV/8202/78607'), ('SO3008929', '1826331', 'INV/8202/77083'), ('SO3008546', '1826331', 'INV/8202/77083'), ('SO3008423', '1826331', 'INV/8202/77083'), ('SO3008108', '1840990', 'INV/8202/78614'), ('SO3000378', '1826328', 'INV/8202/77082'), ('SO3007895', '1826328', 'INV/8202/77082'), ('SO3007719', '1826328', 'INV/8202/77082'), ('SO3007077', '1840990', 'INV/8202/78614'), ('SO3006932', '1826328', 'INV/8202/77082'), ('SO3006856', '1826331', 'INV/8202/77083'), ('SO3006594', '1826340', 'INV/8202/77085'), ('SO3006312', '1826340', 'INV/8202/77085'), ('SO3005999', '1840990', 'INV/8202/78614'), ('SO3005976', '1826328', 'INV/8202/77082'), ('SO3005611', '1826340', 'INV/8202/77085'), ('SO3005159', '1826315', 'INV/8202/77078'), ('SO3005103', '1826340', 'INV/8202/77085'), ('SO3005082', '1826328', 'INV/8202/77082'), ('SO3004876', '1826340', 'INV/8202/77085'), ('SO3004152', '1826340', 'INV/8202/77085'), ('SO3004080', '1826340', 'INV/8202/77085'), ('SO3003236', '1826328', 'INV/8202/77082'), ('SO3002968', '1826328', 'INV/8202/77082'), ('SO3002944', '1841005', 'INV/8202/78619'), ('SO3002496', '1826315', 'INV/8202/77078'), ('SO3002468', '1826347', 'INV/8202/77086'), ('SO3002332', '1826328', 'INV/8202/77082'), ('SO3002329', '1826328', 'INV/8202/77082'), ('SO3002267', '1826328', 'INV/8202/77082'), ('SO3001787', '1826347', 'INV/8202/77086'), ('SO3001755', '1826335', 'INV/8202/77084'), ('SO3001744', '1826347', 'INV/8202/77086'), ('SO3001737', '1837071', 'INV/8202/78195'), ('SO2965482', '1826361', 'INV/8202/77090'), ('SO3000126', '1826361', 'INV/8202/77090'), ('SO2999000', '1826375', 'INV/8202/77092'), ('SO2998643', '1841048', 'INV/8202/78626'), ('SO2998477', '1826335', 'INV/8202/77084'), ('SO2998376', '1826317', 'INV/8202/77079'), ('SO2998320', '1826375', 'INV/8202/77092'), ('SO2998114', '1841048', 'INV/8202/78626'), ('SO2997920', '1826375', 'INV/8202/77092'), ('SO2978604', '1816825', 'INV/8202/76923'), ('SO2968217', '1816825', 'INV/8202/76923'), ('SO2939616', '1816825', 'INV/8202/76923'), ('SO2997567', '1816740', 'INV/8202/76908'), ('SO2997385', '1816825', 'INV/8202/76923'), ('SO2997368', '1816825', 'INV/8202/76923'), ('SO2997334', '1826375', 'INV/8202/77092'), ('SO2997076', '1847119', 'INV/8202/79289'), ('SO2997075', '1816825', 'INV/8202/76923'), ('SO2997042', '1816825', 'INV/8202/76923'), ('SO2996789', '1826317', 'INV/8202/77079'), ('SO2996789', '1826317', 'INV/8202/77079'), ('SO2996760', '1826317', 'INV/8202/77079'), ('SO2996591', '1840867', 'INV/8202/78575'), ('SO2996105', '1816740', 'INV/8202/76908'), ('SO2996035', '1816825', 'INV/8202/76923'), ('SO2996030', '1826378', 'INV/8202/77129'), ('SO2995606', '1816825', 'INV/8202/76923'), ('SO2995571', '1816825', 'INV/8202/76923'), ('SO2995271', '1822284', 'INV/8202/77005'), ('SO2995081', '1826378', 'INV/8202/77129'), ('SO2995019', '1826378', 'INV/8202/77129'), ('SO2994576', '1816825', 'INV/8202/76923'), ('SO2994554', '1816825', 'INV/8202/76923'), ('SO2994453', '1841048', 'INV/8202/78626'), ('SO2994227', '1816831', 'INV/8202/76927'), ('SO2994192', '1816740', 'INV/8202/76908'), ('SO2994159', '1822661', 'INV/8202/77016'), ('SO2994137', '1826378', 'INV/8202/77129'), ('SO2994046', '1826382', 'INV/8202/77093'), ('SO2994046', '1826382', 'INV/8202/77093'), ('SO2994021', '1816831', 'INV/8202/76927'), ('SO2993801', '1816831', 'INV/8202/76927'), ('SO2993719', '1816831', 'INV/8202/76927'), ('SO2993709', '1826378', 'INV/8202/77129'), ('SO2993650', '1816831', 'INV/8202/76927'), ('SO2993538', '1841048', 'INV/8202/78626'), ('SO2945730', '1816740', 'INV/8202/76908'), ('SO2993114', '1816831', 'INV/8202/76927'), ('SO2993059', '1816740', 'INV/8202/76908'), ('SO2993049', '1816831', 'INV/8202/76927'), ('SO2993010', '1826395', 'INV/8202/77095'), ('SO2992998', '1816831', 'INV/8202/76927'), ('SO2992871', '1826395', 'INV/8202/77095'), ('SO2992847', '1826218', 'INV/8202/77064'), ('SO2992466', '1826378', 'INV/8202/77129'), ('SO2992287', '1826395', 'INV/8202/77095'), ('SO2992261', '1816847', 'INV/8202/76928'), ('SO2992117', '1816847', 'INV/8202/76928'), ('SO2991977', '1826378', 'INV/8202/77129'), ('SO2989772', '1816847', 'INV/8202/76928'), ('SO2990557', '1816847', 'INV/8202/76928'), ('SO2991188', '1816847', 'INV/8202/76928'), ('SO2990993', '1816853', 'INV/8202/76929'), ('SO2989925', '1816853', 'INV/8202/76929'), ('SO2989478', '1826395', 'INV/8202/77095'), ('SO2989380', '1826378', 'INV/8202/77129'), ('SO2988985', '1804953', 'INV/8202/76621'), ('SO2988975', '1817016', 'INV/8202/76941'), ('SO2988589', '1841564', 'INV/8202/78750'), ('SO2988574', '1826395', 'INV/8202/77095'), ('SO2987816', '1841564', 'INV/8202/78750'), ('SO2987730', '1817016', 'INV/8202/76941'), ('SO2987648', '1817016', 'INV/8202/76941'), ('SO2987519', '1826393', 'INV/8202/77094'), ('SO2987298', '1817019', 'INV/8202/76943'), ('SO2986187', '1817019', 'INV/8202/76943'), ('SO2985999', '1817019', 'INV/8202/76943'), ('SO2985488', '1817023', 'INV/8202/76944'), ('SO2984990', '1841564', 'INV/8202/78750'), ('SO2984622', '1817023', 'INV/8202/76944'), ('SO2981545', '1826939', 'INV/8202/77124'), ('SO2980929', '1826939', 'INV/8202/77124'), ('SO2980635', '1826939', 'INV/8202/77124'), ('SO2980503', '1817035', 'INV/8202/76948'), ('SO2980478', '1826393', 'INV/8202/77094'), ('SO2980068', '1817035', 'INV/8202/76948'), ('SO2979826', '1817035', 'INV/8202/76948'), ('SO2978778', '1817040', 'INV/8202/76951'), ('SO2978719', '1826393', 'INV/8202/77094'), ('SO2978514', '1817040', 'INV/8202/76951'), ('SO2978168', '1816755', 'INV/8202/76911'), ('SO2977587', '1826393', 'INV/8202/77094'), ('SO2976760', '1827077', 'INV/8202/77140'), ('SO2976743', '1841733', 'INV/8202/78791'), ('SO2975583', '1841061', 'INV/8202/78666'), ('SO2974964', '1841061', 'INV/8202/78666'), ('SO2974747', '1841733', 'INV/8202/78791'), ('SO2974405', '1816770', 'INV/8202/76934'), ('SO2974272', '1817072', 'INV/8202/76954'), ('SO2974092', '1817072', 'INV/8202/76954'), ('SO2973802', '1803403', 'INV/8202/76557'), ('SO2972841', '1827103', 'INV/8202/77145'), ('SO2972573', '1841733', 'INV/8202/78791'), ('SO2972461', '1841061', 'INV/8202/78666'), ('SO2972140', '1817106', 'INV/8202/76956'), ('SO2971925', '1805152', 'INV/8202/76643'), ('SO2963065', '1784543', 'INV/8202/76286'), ('SO2971786', '1827103', 'INV/8202/77145'), ('SO2971742', '1827103', 'INV/8202/77145'), ('SO2971727', '1827103', 'INV/8202/77145'), ('SO2971102', '1817112', 'INV/8202/76968'), ('SO2971035', '1806791', 'INV/8202/76773'), ('SO2956464', '1816770', 'INV/8202/76934'), ('SO2970897', '1826393', 'INV/8202/77094'), ('SO2970897', '1826393', 'INV/8202/77094'), ('SO2969993', '1847119', 'INV/8202/79289'), ('SO2969648', '1817133', 'INV/8202/76958'), ('SO2969126', '1817133', 'INV/8202/76958'), ('SO2968632', '1847119', 'INV/8202/79289'), ('SO2968137', '1782166', 'INV/8202/76254'), ('SO2968044', '1817147', 'INV/8202/76960'), ('SO2967108', '1826324', 'INV/8202/77081'), ('SO2967106', '1826324', 'INV/8202/77081'), ('SO2965608', '1817158', 'INV/8202/76962'), ('SO2964568', '1817143', 'INV/8202/76959'), ('SO2963595', '1817143', 'INV/8202/76959'), ('SO2963329', '1817143', 'INV/8202/76959'), ('SO2963085', '1816770', 'INV/8202/76934'), ('SO2963026', '1817175', 'INV/8202/76963'), ('SO2961644', '1805299', 'INV/8202/76666'), ('SO2961621', '1816783', 'INV/8202/76935'), ('SO2961100', '1805751', 'INV/8202/76745'), ('SO2960908', '1817175', 'INV/8202/76963'), ('SO2960734', '1805348', 'INV/8202/76671'), ('SO2960475', '1816783', 'INV/8202/76935'), ('SO2960318', '1805348', 'INV/8202/76671'), ('SO2960315', '1805348', 'INV/8202/76671'), ('SO2960306', '1805348', 'INV/8202/76671'), ('SO2960305', '1805348', 'INV/8202/76671'), ('SO2960302', '1805348', 'INV/8202/76671'), ('SO2960250', '1805751', 'INV/8202/76745'), ('SO2959971', '1805348', 'INV/8202/76671'), ('SO2959693', '1817175', 'INV/8202/76963'), ('SO2959439', '1817175', 'INV/8202/76963'), ('SO2958707', '1817175', 'INV/8202/76963'), ('SO2957401', '1805811', 'INV/8202/76747'), ('SO2957328', '1841733', 'INV/8202/78791'), ('SO2957328', '1841733', 'INV/8202/78791'), ('SO2956970', '1805811', 'INV/8202/76747'), ('SO2956880', '1805435', 'INV/8202/76682'), ('SO2956340', '1797066', 'INV/8202/76413'), ('SO2955847', '1841733', 'INV/8202/78791'), ('SO2955847', '1841733', 'INV/8202/78791'), ('SO2954412', '1771383', 'INV/8202/75980'), ('SO2952160', '1805590', 'INV/8202/76717'), ('SO2951969', '1805603', 'INV/8202/76718'), ('SO2950948', '1773466', 'INV/8202/76021'), ('SO2950183', '1774336', 'INV/8202/76082'), ('SO2949968', '1800919', 'INV/8202/76469'), ('SO2949532', '1811473', 'INV/8202/76804'), ('SO2948953', '1805640', 'INV/8202/76723'), ('SO2948047', '1841564', 'INV/8202/78750'), ('SO2948047', '1841564', 'INV/8202/78750'), ('SO2946770', '1805669', 'INV/8202/76727'), ('SO2944673', '1805285', 'INV/8202/76939'), ('SO2944528', '1810573', 'INV/8202/76789'), ('SO2944457', '1811473', 'INV/8202/76804'), ('SO2939973', '1805624', 'INV/8202/76719'), ('SO2938775', '1754370', 'INV/8202/75543'), ('SO2938153', '1805647', 'INV/8202/76724'), ('SO2937233', '1840900', 'INV/8202/78587'), ('SO2937226', '1767223', 'INV/8202/75921'), ('SO2936947', '1805381', 'INV/8202/76674'), ('SO2935185', '1841733', 'INV/8202/78791'), ('SO2935185', '1841733', 'INV/8202/78791'), ('SO2934012', '1805442', 'INV/8202/76683'), ('SO2926614', '1805559', 'INV/8202/76713'), ('SO2925673', '1841564', 'INV/8202/78750'), ('SO2925673', '1841564', 'INV/8202/78750'), ('SO2924721', '1804205', 'INV/8202/76563'), ('SO2923702', '1810732', 'INV/8202/76793'), ('SO2923259', '1841556', 'INV/8202/78745'), ('SO2923259', '1841556', 'INV/8202/78745'), ('SO2917492', '1764711', 'INV/8202/75888'), ('SO2905437', '1816792', 'INV/8202/76916'), ('SO2903468', '1754254', 'INV/8202/75526')]
    #Lista de SO a las que se les creó una credit_notes
    so_modified = []
    #Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    #Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    #Lista de nombres de las notas de crédito creadas
    nc_created = []
    #Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    #Lista de facturas origen
    so_origin_invoice = []
    #Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    #Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0] # Almacena el nombre de la SO
            inv_id = each[1] # Almacena el ID de la factura
            inv_name = each[2] # Almacena el nombre de la factura
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_usage = 'G02'  # Uso del CFDI
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_uuid_origin = f'03|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    inv_payment = inv['l10n_mx_edi_payment_method_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        #--------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        #Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                        if not existing_credit_note:
                            try:
                                #Busca la órden de venta
                                sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read', [[['name', '=', inv_origin_name]]])[0]
                                # Obtiene los datos necesarios directo de la SO
                                sale_id = sale_order['id']
                                sale_name = sale_order['name']
                                sale_ref = sale_order['channel_order_reference']
                                sale_team = sale_order['team_id'][0]
                                #Busca el order line correspondiente de la orden de venta
                                sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line', 'search_read', [[['order_id', '=', sale_id]]])
                                #Define los valores de la nota de crédito
                                inv_int = int(inv_id)
                                sale_int = int(sale_id)
                                refund_vals = {
                                    'ref': f'Reversión de: {inv_name}',
                                    'journal_id': inv_journal_id,
                                    'team_id': sale_team,
                                    'invoice_origin': sale_name,
                                    'payment_reference': inv_name,
                                    'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                    # Puedes ajustar la fecha según tus necesidades
                                    'partner_id': inv['partner_id'][0],
                                    'l10n_mx_edi_usage': inv_usage,
                                    'l10n_mx_edi_origin': inv_uuid_origin,
                                    'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id,
                                    'reversed_entry_id': inv_int,
                                    'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                    'invoice_line_ids': []
                                }
                                for lines in sale_line_id:
                                    nc_lines = {'product_id': lines['product_id'][0],
                                                'quantity': lines['product_uom_qty'],
                                                'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                                'price_unit': lines['price_unit'],
                                                'product_uom_id': lines['product_uom'][0],
                                                'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                                }
                                    refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                                #Crea la nota de crédito
                                create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create', [refund_vals])
                                #Actualiza la nota de crédito
                                #Agrega mensaje al Attachment de la nota de crédito
                                message = {
                                    'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                    'message_type': 'comment',
                                }
                                write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[create_nc], message)
                                #Enlazamos la venta con la nueva factura
                                upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write', [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                                #Publicamos la nota de crédito
                                #upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                                #Timbramos la nota de crédito
                                #upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                                #Buscamos el nombre de la factura ya creada
                                search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', create_nc]]])
                                nc_name = search_nc_name[0]['name']
                                nc_total = search_nc_name[0]['amount_total']
                                #Agregamos a las listas
                                so_modified.append(sale_name)
                                nc_created.append(nc_name)
                                nc_amount_total.append(nc_total)
                                so_origin_invoice.append(inv_name)
                                so_mkp_reference.append(sale_ref)
                                progress_bar.update(1)
                            except Exception as b:
                                print(f"En el armado de la factura y la creación: {b}")
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    #Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_globales_meli_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['sergio@wonderbrands.co', 'eric@wonderbrands.co','rosalba@wonderbrands.co', 'natalia@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Meli - Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC globales Meli completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()

def reverse_invoice_amazon():
    # Formato para query
    type_filter = 'INDIVIDUAL'
    marketplace_filter = 'AMAZON'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_amz_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO INDIVIDUALES AMAZON')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    #INDIVIDUALES AMAZON
    mycursor.execute("""SELECT c.name,
                               b.id 'account_move_id',
                               d.refund_date as 'payment_date_last_modified'/*,
                               d.order_id 'order_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               d.refunded_amt,
                               b.invoice_partner_display_name 'cliente',
                               b.name,
                               'INDIVIDUAL' as type,
                               'AMAZON' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON b.invoice_origin = c.name
                        LEFT JOIN (SELECT a.order_id, max(STR_TO_DATE(fecha, '%d/%m/%Y')) 'refund_date', SUM(total - tarifas_de_amazon) * (-1) 'refunded_amt'
                                   FROM somos_reyes.amazon_payments_refunds a
                                   WHERE (total - tarifas_de_amazon) * (-1) > 0 AND STR_TO_DATE(fecha, '%d/%m/%Y') >= %s AND STR_TO_DATE(fecha, '%d/%m/%Y') <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE d.order_id is not null
                        AND e.invoice_origin is null
                        AND d.refunded_amt - b.amount_total < 1 AND d.refunded_amt - b.amount_total > -1
                        AND c.name in ({});""".format(placeholders), tuple(dates_list_params+list_orders))

    invoice_records = mycursor.fetchall()
    # Lista de SO a las que se les creó una credit_notes
    so_modified = []
    # Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    # Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    # Lista de nombres de las notas de crédito creadas
    nc_created = []
    # Lista de facturas origen
    so_origin_invoice = []
    # Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    # Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0]
            inv_id = each[1]
            nc_date = each[2].strftime("%Y-%m-%d %H:%M:%S")
            inv_move_types = [] # Lista en la que se almacenan los tipos de factura para la orden en curso
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['invoice_origin', '=', inv_origin_name]]])
            if invoice:
                for type in invoice:
                    exist_nc_type = type['move_type']
                    inv_move_types.append(exist_nc_type)

                # Se verifica si ya existe una nota de crédito para esta orden de venta
                existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                if not existing_credit_note:
                #if 'out_refund' not in inv_move_types:
                    for inv in invoice:
                        inv_id = inv['id'] # ID de la factura
                        inv_name = inv['name'] # Nombre de la factura
                        inv_origin = inv['invoice_origin'] # Nombre de la SO ligada a la factura
                        #inv_narration = inv['narration']
                        #inv_uuid = inv_narration[3:-4]
                        inv_uuid = inv['l10n_mx_edi_cfdi_uuid'] # Folio fiscal de la factura
                        inv_journal_id = inv['journal_id'][0] #Diario de la factura
                        l10n_mx_edi_origin = '03|' + str(inv_uuid)
                        team_id = inv['team_id'][0]
                        #Se hace una llamada al wizard de creación de notas de crédito
                        credit_note_wizard = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'create',
                                                               [{
                            'refund_method': 'refund',
                            'reason': 'Por efectos de devolución o retorno de una orden',
                            'journal_id': inv_journal_id, }],
                                       {'context': {
                                           'active_ids': [inv_id],
                                           'active_id': inv_id,
                                           'active_model': 'account.move',
                                       }}
                                    )
                        #Se crea la nota de crédito con la info anterior y se usa la función reverse_moves del botón revertir en el wizard
                        nc_inv_create = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'reverse_moves',[credit_note_wizard])
                        nc_id = nc_inv_create['res_id'] # Obtiene el id de la nota de crédito
                        # Agrega un mensaje al chatter de la nota de crédito
                        message = {
                            'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {inv_origin}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                            'message_type': 'comment',
                        }
                        write_msg_tech = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[nc_id], message)
                        # Actualización de Forma de Pago, CFDI Origen, Equipo de Ventas
                        update_vals_nc = models.execute_kw(db_name, uid, password, 'account.move', 'write', [[nc_id], {'team_id': team_id, 'l10n_mx_edi_origin': l10n_mx_edi_origin, 'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id, 'l10n_mx_edi_usage': l10n_mx_edi_usage}])
                        #Confirma la nota de crédito
                        #upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post',[nc_id])
                        # Timbramos la nota de crédito
                        # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[nc_id])
                        #buscamos el nombre de la nota creada
                        search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', nc_id]]])
                        nc_name = search_nc_name[0]['name']
                        nc_total = search_nc_name[0]['amount_total']
                        sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read',[[['name', '=', inv_origin_name]]])[0]
                        sale_ref = sale_order['channel_order_reference']
                        #Agregamos a las listas
                        nc_created.append(nc_name)
                        nc_amount_total.append(nc_total)
                        so_modified.append(inv_origin)
                        so_origin_invoice.append(inv_name)
                        so_mkp_reference.append(sale_ref)
                        progress_bar.update(1)
                else:
                    print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                    so_with_refund.append(inv_origin_name)
                    progress_bar.update(1)
                    continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_individuales_amazon_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['sergio@wonderbrands.co', 'eric@wonderbrands.co', 'rosalba@wonderbrands.co', 'natalia@wonderbrands.co', 'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Amazon - Creación de notas de crédito para facturas individuales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('PROCESO NC PARA AMAZON COMPLETADO :)')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
def reverse_invoice_global_amazon():
    # Formato para query
    type_filter = 'GLOBAL'
    marketplace_filter = 'AMAZON'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_amz_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO GLOBALES AMAZON')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')

    mycursor.execute("""#GLOBALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name/*,
                               d.order_id,
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               d.refunded_amt,
                               refund_date,
                               b.invoice_partner_display_name 'cliente',
                               'GLOBAL' as type,
                               'AMAZON' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON SUBSTRING_INDEX(SUBSTRING_INDEX(invoice_ids, ']', 1), '[', -1) = b.id
                        LEFT JOIN (SELECT a.order_id, max(STR_TO_DATE(fecha, '%d/%m/%Y')) 'refund_date', SUM(total - tarifas_de_amazon) * (-1) 'refunded_amt'
                                   FROM somos_reyes.amazon_payments_refunds a
                                   WHERE (total - tarifas_de_amazon) * (-1) > 0 AND STR_TO_DATE(fecha, '%d/%m/%Y') >= %s AND STR_TO_DATE(fecha, '%d/%m/%Y') <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE d.order_id is not null
                        AND e.invoice_origin is null
                        AND invoice_partner_display_name = 'PÚBLICO EN GENERAL'
                        AND (d.refunded_amt - b.amount_total > 1 OR d.refunded_amt - b.amount_total < -1)
                        AND d.refunded_amt - c.amount_total < 1 AND d.refunded_amt - c.amount_total > -1
                        AND c.name in ({});
                        """.format(placeholders), tuple(dates_list_params+list_orders))
    invoice_records = mycursor.fetchall()
    #Lista de SO a las que se les creó una credit_notes
    so_modified = []
    #Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    #Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    #Lista de nombres de las notas de crédito creadas
    nc_created = []
    #Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    #Lista de facturas origen
    so_origin_invoice = []
    #Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    #Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0] # Almacena el nombre de la SO
            inv_id = each[1] # Almacena el ID de la factura
            inv_name = each[2] # Almacena el nombre de la factura
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_usage = inv['l10n_mx_edi_usage']  # Folio fiscal de la factura
                    inv_uuid_origin = f'03|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        #--------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        #Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                        if not existing_credit_note:
                            #Busca la órden de venta
                            sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read', [[['name', '=', inv_origin_name]]])[0]
                            # Obtiene los datos necesarios directo de la SO
                            sale_id = sale_order['id']
                            sale_name = sale_order['name']
                            sale_ref = sale_order['channel_order_reference']
                            sale_team = sale_order['team_id'][0]
                            #Busca el order line correspondiente de la orden de venta
                            sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line', 'search_read', [[['order_id', '=', sale_id]]])
                            #Define los valores de la nota de crédito
                            inv_int = int(inv_id)
                            sale_int = int(sale_id)
                            refund_vals = {
                                'ref': f'Reversión de: {inv_name}',
                                'journal_id': inv_journal_id,
                                'invoice_origin': sale_name,
                                'team_id': sale_team,
                                'payment_reference': inv_name,
                                'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                # Puedes ajustar la fecha según tus necesidades
                                'partner_id': inv['partner_id'][0],
                                'l10n_mx_edi_usage': l10n_mx_edi_usage,
                                'l10n_mx_edi_origin': inv_uuid_origin,
                                'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id,
                                'reversed_entry_id': inv_int,
                                'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                'invoice_line_ids': []
                            }
                            for lines in sale_line_id:
                                nc_lines = {'product_id': lines['product_id'][0],
                                            'quantity': lines['product_uom_qty'],
                                            'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                            'price_unit': lines['price_unit'],
                                            'product_uom_id': lines['product_uom'][0],
                                            'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                            }
                                refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                            #Crea la nota de crédito
                            create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create', [refund_vals])
                            #Actualiza la nota de crédito
                            #Agrega mensaje al Attachment de la nota de crédito
                            message = {
                                'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                'message_type': 'comment',
                            }
                            write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[create_nc], message)
                            #Enlazamos la venta con la nueva factura
                            upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write', [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                            #Publicamos la nota de crédito
                            #upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                            #Timbramos la nota de crédito
                            #upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                            #Buscamos el nombre de la factura ya creada
                            search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', create_nc]]])
                            nc_name = search_nc_name[0]['name']
                            nc_total = search_nc_name[0]['amount_total']
                            #Agregamos a las listas
                            so_modified.append(sale_name)
                            nc_created.append(nc_name)
                            nc_amount_total.append(nc_total)
                            so_origin_invoice.append(inv_name)
                            so_mkp_reference.append(sale_ref)
                            progress_bar.update(1)
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
       print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    #Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_globales_amazon_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['sergio@wonderbrands.co', 'eric@wonderbrands.co', 'rosalba@wonderbrands.co', 'natalia@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Amazon - Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC Amazon completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()

if __name__ == "__main__":
    reverse_invoice_meli()
    reverse_invoice_global_meli()
    reverse_invoice_amazon()
    reverse_invoice_global_amazon()
    end_time = datetime.datetime.now()
    duration = end_time - today_date
    print(f'Duraciòn del script: {duration}')
    print('Listo')
    print('Este arroz ya se coció :)')