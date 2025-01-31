from email.message import EmailMessage
from email.utils import make_msgid
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from pprint import pprint
from email import encoders
from tqdm import tqdm
import time
import json
# import jsonrpc
# import jsonrpclib
# import random
# import urllib.request
# import getpass
# import http
# import requests
# import logging
# import zipfile
# import socket
import os
import locale
import xmlrpc.client
import base64
import openpyxl
# import xlrd
# import pandas as pd
# import MySQLdb
import mysql.connector
import smtplib
# import ssl
# import email
import datetime
import time
from concurrent.futures import ThreadPoolExecutor

from Test import extract_orders as e_o
from Test import prepare_folders as prep

print('================================================================')
print('BIENVENIDO AL PROCESO DE NOTAS DE CRÉDITO PARA MARKETPLACES')
print('================================================================')
print('SCRIPT DE CREACIÓN DE NOTAS DE CRÉDITO')
print('================================================================')
today_date = datetime.datetime.now()
dir_path = os.path.dirname(os.path.realpath(__file__))
print('Fecha:' + today_date.strftime("%Y-%m-%d %H:%M:%S"))
#Archivo de configuración - Use config_dev.json si está haciendo pruebas
#Archivo de configuración - Use config.json cuando los cambios vayan a producción

# ***********************************************
# ARCHIVO DE CONFIGURACIÓN
config_file = 'config.json'
# ***********************************************

config_file_name = rf'C:\Users\Sergio Gil Guerrero\Documents\WonderBrands\Repos\wb_odoo_external_api\config\{config_file}'
l10n_mx_edi_payment_method_id = 3
l10n_mx_edi_usage = 'G02'

#FECHAS DEL PERIODO ***********************************************
start_date_str = datetime.date(2024, 12, 27).strftime("%Y-%m-%d")
end_date_str = datetime.date(2025, 1, 28).strftime("%Y-%m-%d")
# ***********************************************

month_executed, year_executed = prep.get_dates()
year_executed = str(year_executed)

# ----------------------------------------------------------------
# Mes y año manual
# month_executed = 'Septiembre'
# year_executed = '2024'
# ----------------------------------------------------------------


print('******************************************')
print(month_executed, year_executed)
print('******************************************')

#PATHS de los archivos de ordenes conciliadas
orders_meli_file_path = 'C:/Users/Sergio Gil Guerrero/Documents/WonderBrands/Finanzas/{}/{}/Conciliadas/Notas_de_credito_totales_ML.csv'.format(year_executed,month_executed)
orders_amz_file_path = 'C:/Users/Sergio Gil Guerrero/Documents/WonderBrands/Finanzas/{}/{}/Conciliadas/Notas_de_credito_totales_AMZ.csv'.format(year_executed,month_executed)

def get_odoo_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['odoo']
def get_psql_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['psql']
def get_email_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['email']

def current_execution(func):
    def wrapper(*args, **kwargs):
        print('\n \n ******************************************************')
        print(f"Ejecutando {func.__name__}")
        result = func(*args, **kwargs)
        print(f"{func.__name__} terminada")
        print('****************************************************** \n \n')
        return result
    return wrapper

@current_execution
def reverse_invoice_meli(): #NOTAS DE CRÉDITO INDIVIDUALES MELI
    #Formato para query
    type_filter = 'INDIVIDUAL'
    marketplace_filter = 'MERCADO LIBRE'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_meli_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str, start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO INDIVIDUALES MELI')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    #INDIVIDUALES MELI
    mycursor.execute("""#INDIVIDUALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               ifnull(d.payment_date_last_modified, dd.payment_date_last_modified) 'payment_date_last_modified'/*,
                               ifnull(d.order_id, dd.pack_id) 'order_id_or_pack_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               ifnull(d.refunded_amt, dd.refunded_amt) 'ml_refunded_amount',
                               b.invoice_partner_display_name 'cliente',
                               b.name,
                               'INDIVIDUAL' as type,
                               'MERCADO LIBRE' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON b.invoice_origin = c.name
                        LEFT JOIN (SELECT a.order_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(paid_amt) 'paid_amt', SUM(refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                                   FROM somos_reyes.ml_order_payments a
                                   LEFT JOIN somos_reyes.ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT a.pack_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(b.paid_amt) 'paid_amt', SUM(b.refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                        FROM somos_reyes.ml_order_update a
                        LEFT JOIN somos_reyes.ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                        GROUP BY 1) dd
                        ON c.yuju_pack_id = dd.pack_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE (d.order_id is not null or dd.pack_id is not null)
                        AND e.invoice_origin is null
                        AND ((ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total < 1 AND ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total > -1)
                        OR (ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total < 1
                        AND ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total > -1))
                        AND c.name in ({});
                            """.format(placeholders), tuple(dates_list_params+list_orders))
    invoice_records = mycursor.fetchall()
    # Lista de SO a las que se les creó una credit_notes
    so_modified = []
    # Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    # Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    # Lista de nombres de las notas de crédito creadas
    nc_created = []
    # Lista de facturas origen
    so_origin_invoice = []
    # Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    # Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0] #Es la SO
            inv_id = each[1]
            nc_date = each[2].strftime("%Y-%m-%d %H:%M:%S")
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['invoice_origin', '=', inv_origin_name]]])
            if invoice:
                # Se verifica si ya existe una nota de crédito para esta orden de venta
                existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                if not existing_credit_note:
                #if 'out_refund' not in inv_move_types:
                    for inv in invoice:
                        inv_id = inv['id'] # ID de la factura
                        inv_name = inv['name'] # Nombre de la factura
                        inv_origin = inv['invoice_origin'] # Nombre de la SO ligada a la factura
                        #inv_narration = inv['narration']
                        #inv_uuid = inv_narration[3:-4]
                        inv_uuid = inv['l10n_mx_edi_cfdi_uuid'] # Folio fiscal de la factura
                        inv_journal_id = inv['journal_id'][0] #Diario de la factura
                        inv_state = inv['state']
                        l10n_mx_edi_origin = '03|' + str(inv_uuid)
                        team_id = inv['team_id'][0]

                        if inv_state == 'posted':
                            #Se hace una llamada al wizard de creación de notas de crédito
                            credit_note_wizard = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'create',
                                                                   [{
                                'refund_method': 'refund',
                                'reason': 'Por efectos de devolución o retorno de una orden',
                                'journal_id': inv_journal_id, }],
                                           {'context': {
                                               'active_ids': [inv_id],
                                               'active_id': inv_id,
                                               'active_model': 'account.move',
                                           }}
                                        )
                            #Se crea la nota de crédito con la info anterior y se usa la función reverse_moves del botón revertir en el wizard
                            nc_inv_create = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'reverse_moves',[credit_note_wizard])
                            nc_id = nc_inv_create['res_id'] # Obtiene el id de la nota de crédito
                            # Agrega un mensaje al chatter de la nota de crédito
                            message = {
                                'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {inv_origin}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                'message_type': 'comment',
                            }
                            write_msg_tech = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[nc_id], message)
                            #Actualización de Forma de Pago, CFDI Origen, Equipo de Ventas
                            update_vals_nc = models.execute_kw(db_name, uid, password, 'account.move', 'write',[[nc_id], {'team_id': team_id, 'l10n_mx_edi_origin': l10n_mx_edi_origin, 'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id, 'l10n_mx_edi_usage': l10n_mx_edi_usage}])
                            #Confirma la nota de crédito
                            upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post',[nc_id])
                            # Timbramos la nota de crédito
                            # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[nc_id])
                            stamp_credit_note(models, db_name, uid, password, nc_id)

                            #buscamos el nombre de la nota creada
                            search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', nc_id]]])
                            nc_name = search_nc_name[0]['name']
                            nc_total = search_nc_name[0]['amount_total']
                            sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read',[[['name', '=', inv_origin_name]]])[0]
                            sale_ref = sale_order['channel_order_reference']
                            #Agregamos a las listas
                            nc_created.append(nc_name)
                            nc_amount_total.append(nc_total)
                            so_modified.append(inv_origin)
                            so_origin_invoice.append(inv_name)
                            so_mkp_reference.append(sale_ref)
                            progress_bar.update(1)
                        else:
                            print(f"La factura de la órden {inv_origin_name} no está confirmada")
                            progress_bar.update(1)
                            continue
                else:
                    print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                    so_with_refund.append(inv_origin_name)
                    progress_bar.update(1)
                    continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_individuales_meli_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales 
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito 
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron 
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['carlos.hinojosa@wonderbrands.co', 'sergio@wonderbrands.co', 'eric@wonderbrands.co','rosalba@wonderbrands.co', 'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Meli - Creación de notas de crédito para facturas individuales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('PROCESO NC PARA MERCADO LIBRE COMPLETADO :)')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
@current_execution
def reverse_invoice_global_meli():
    # Formato para query
    type_filter = 'GLOBAL'
    marketplace_filter = 'MERCADO LIBRE'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_meli_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str, start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO GLOBALES MELI')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    #GLOBALES MELI
    mycursor.execute("""#GLOBALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name/*,
                               ifnull(d.order_id, dd.pack_id) 'order_id_or_pack_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               ifnull(d.refunded_amt, dd.refunded_amt) 'ml_refunded_amount',
                               ifnull(d.payment_date_last_modified, dd.payment_date_last_modified) 'payment_date_last_modified',
                               b.invoice_partner_display_name 'cliente',
                               'GLOBAL' as type,
                               'MERCADO LIBRE' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON SUBSTRING_INDEX(SUBSTRING_INDEX(invoice_ids, ']', 1), '[', -1) = b.id
                        LEFT JOIN (SELECT a.order_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(paid_amt) 'paid_amt', SUM(refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                                   FROM somos_reyes.ml_order_payments a
                                   LEFT JOIN somos_reyes.ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT a.pack_id, max(payment_date_last_modified) 'payment_date_last_modified', SUM(b.paid_amt) 'paid_amt', SUM(b.refunded_amt) 'refunded_amt', SUM(shipping_amt) 'shipping_amt'
                        FROM somos_reyes.ml_order_update a
                        LEFT JOIN somos_reyes.ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= %s AND date(payment_date_last_modified) <= %s
                        GROUP BY 1) dd
                        ON c.yuju_pack_id = dd.pack_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE (d.order_id is not null or dd.pack_id is not null)
                        AND e.invoice_origin is null
                        AND invoice_partner_display_name = 'PÚBLICO EN GENERAL'
                        AND (ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total > 1 OR ifnull(d.refunded_amt, dd.refunded_amt) - b.amount_total < -1)
                        AND (ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total > 1 OR ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - b.amount_total < -1)
                        AND ((ifnull(d.refunded_amt, dd.refunded_amt) - c.amount_total < 1 AND ifnull(d.refunded_amt, dd.refunded_amt) - c.amount_total > -1)
                        OR (ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - c.amount_total < 1
                        AND ifnull(d.refunded_amt - d.shipping_amt, dd.refunded_amt - dd.shipping_amt) - c.amount_total > -1))
                        AND c.name in ({});""".format(placeholders), tuple(dates_list_params+list_orders))
    invoice_records = mycursor.fetchall()
    #Lista de SO a las que se les creó una credit_notes
    so_modified = []
    #Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    #Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    #Lista de nombres de las notas de crédito creadas
    nc_created = []
    #Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    #Lista de facturas origen
    so_origin_invoice = []
    #Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    #Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0] # Almacena el nombre de la SO
            inv_id = each[1] # Almacena el ID de la factura
            inv_name = each[2] # Almacena el nombre de la factura
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_usage = 'G02'  # Uso del CFDI
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_uuid_origin = f'03|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    inv_payment = inv['l10n_mx_edi_payment_method_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        #--------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        #Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                        if not existing_credit_note:
                            try:
                                #Busca la órden de venta
                                sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read', [[['name', '=', inv_origin_name]]])[0]
                                # Obtiene los datos necesarios directo de la SO
                                sale_id = sale_order['id']
                                sale_name = sale_order['name']
                                sale_ref = sale_order['channel_order_reference']
                                sale_team = sale_order['team_id'][0]
                                #Busca el order line correspondiente de la orden de venta
                                sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line', 'search_read', [[['order_id', '=', sale_id]]])
                                #Define los valores de la nota de crédito
                                inv_int = int(inv_id)
                                sale_int = int(sale_id)
                                refund_vals = {
                                    'ref': f'Reversión de: {inv_name}',
                                    'journal_id': inv_journal_id,
                                    'team_id': sale_team,
                                    'invoice_origin': sale_name,
                                    'payment_reference': inv_name,
                                    'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                    # Puedes ajustar la fecha según tus necesidades
                                    'partner_id': inv['partner_id'][0],
                                    'l10n_mx_edi_usage': inv_usage,
                                    'l10n_mx_edi_origin': inv_uuid_origin,
                                    'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id,
                                    'reversed_entry_id': inv_int,
                                    'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                    'invoice_line_ids': []
                                }
                                for lines in sale_line_id:
                                    nc_lines = {'product_id': lines['product_id'][0],
                                                'quantity': lines['product_uom_qty'],
                                                'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                                'price_unit': lines['price_unit'],
                                                'product_uom_id': lines['product_uom'][0],
                                                'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                                }
                                    refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                                #Crea la nota de crédito
                                create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create', [refund_vals])
                                #Actualiza la nota de crédito
                                #Agrega mensaje al Attachment de la nota de crédito
                                message = {
                                    'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                    'message_type': 'comment',
                                }
                                write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[create_nc], message)
                                #Enlazamos la venta con la nueva factura
                                upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write', [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                                #Publicamos la nota de crédito
                                upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                                #Timbramos la nota de crédito
                                #upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                                stamp_credit_note(models, db_name, uid, password, create_nc)

                                #Buscamos el nombre de la factura ya creada
                                search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', create_nc]]])
                                nc_name = search_nc_name[0]['name']
                                nc_total = search_nc_name[0]['amount_total']
                                #Agregamos a las listas
                                so_modified.append(sale_name)
                                nc_created.append(nc_name)
                                nc_amount_total.append(nc_total)
                                so_origin_invoice.append(inv_name)
                                so_mkp_reference.append(sale_ref)
                                progress_bar.update(1)
                            except Exception as b:
                                print(f"En el armado de la factura y la creación: {b}")
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
       print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    #Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_globales_meli_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['carlos.hinojosa@wonderbrands.co', 'sergio@wonderbrands.co', 'eric@wonderbrands.co','rosalba@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Meli - Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC globales Meli completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
@current_execution
def reverse_invoice_amazon():
    # Formato para query
    type_filter = 'INDIVIDUAL'
    marketplace_filter = 'AMAZON'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_amz_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO INDIVIDUALES AMAZON')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    #INDIVIDUALES AMAZON
    mycursor.execute("""SELECT c.name,
                               b.id 'account_move_id',
                               d.refund_date as 'payment_date_last_modified'/*,
                               d.order_id 'order_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               d.refunded_amt,
                               b.invoice_partner_display_name 'cliente',
                               b.name,
                               'INDIVIDUAL' as type,
                               'AMAZON' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON b.invoice_origin = c.name
                        LEFT JOIN (SELECT a.order_id, max(STR_TO_DATE(fecha, '%m/%d/%Y')) 'refund_date', SUM(total - tarifas_de_amazon) * (-1) 'refunded_amt'
                                   FROM somos_reyes.amazon_payments_refunds a
                                   WHERE (total - tarifas_de_amazon) * (-1) > 0 AND STR_TO_DATE(fecha, '%m/%d/%Y') >= %s AND STR_TO_DATE(fecha, '%m/%d/%Y') <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE d.order_id is not null
                        AND e.invoice_origin is null
                        AND d.refunded_amt - b.amount_total < 1 AND d.refunded_amt - b.amount_total > -1
                        AND c.name in ({});""".format(placeholders), tuple(dates_list_params+list_orders))

    invoice_records = mycursor.fetchall()
    # Lista de SO a las que se les creó una credit_notes
    so_modified = []
    # Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    # Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    # Lista de nombres de las notas de crédito creadas
    nc_created = []
    # Lista de facturas origen
    so_origin_invoice = []
    # Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    # Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0]
            inv_id = each[1]
            nc_date = each[2].strftime("%Y-%m-%d %H:%M:%S")
            inv_move_types = [] # Lista en la que se almacenan los tipos de factura para la orden en curso
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['invoice_origin', '=', inv_origin_name]]])
            if invoice:
                for type in invoice:
                    exist_nc_type = type['move_type']
                    inv_move_types.append(exist_nc_type)

                # Se verifica si ya existe una nota de crédito para esta orden de venta
                existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                if not existing_credit_note:
                #if 'out_refund' not in inv_move_types:
                    for inv in invoice:
                        inv_id = inv['id'] # ID de la factura
                        inv_name = inv['name'] # Nombre de la factura
                        inv_origin = inv['invoice_origin'] # Nombre de la SO ligada a la factura
                        #inv_narration = inv['narration']
                        #inv_uuid = inv_narration[3:-4]
                        inv_uuid = inv['l10n_mx_edi_cfdi_uuid'] # Folio fiscal de la factura
                        inv_journal_id = inv['journal_id'][0] #Diario de la factura
                        l10n_mx_edi_origin = '03|' + str(inv_uuid)
                        team_id = inv['team_id'][0]
                        #Se hace una llamada al wizard de creación de notas de crédito
                        credit_note_wizard = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'create',
                                                               [{
                            'refund_method': 'refund',
                            'reason': 'Por efectos de devolución o retorno de una orden',
                            'journal_id': inv_journal_id, }],
                                       {'context': {
                                           'active_ids': [inv_id],
                                           'active_id': inv_id,
                                           'active_model': 'account.move',
                                       }}
                                    )
                        #Se crea la nota de crédito con la info anterior y se usa la función reverse_moves del botón revertir en el wizard
                        nc_inv_create = models.execute_kw(db_name, uid, password, 'account.move.reversal', 'reverse_moves',[credit_note_wizard])
                        nc_id = nc_inv_create['res_id'] # Obtiene el id de la nota de crédito
                        # Agrega un mensaje al chatter de la nota de crédito
                        message = {
                            'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {inv_origin}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                            'message_type': 'comment',
                        }
                        write_msg_tech = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[nc_id], message)
                        # Actualización de Forma de Pago, CFDI Origen, Equipo de Ventas
                        update_vals_nc = models.execute_kw(db_name, uid, password, 'account.move', 'write', [[nc_id], {'team_id': team_id, 'l10n_mx_edi_origin': l10n_mx_edi_origin, 'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id, 'l10n_mx_edi_usage': l10n_mx_edi_usage}])
                        #Confirma la nota de crédito
                        upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post',[nc_id])
                        # Timbramos la nota de crédito
                        # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[nc_id])
                        stamp_credit_note(models, db_name, uid, password, nc_id)

                        #buscamos el nombre de la nota creada
                        search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', nc_id]]])
                        nc_name = search_nc_name[0]['name']
                        nc_total = search_nc_name[0]['amount_total']
                        sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read',[[['name', '=', inv_origin_name]]])[0]
                        sale_ref = sale_order['channel_order_reference']
                        #Agregamos a las listas
                        nc_created.append(nc_name)
                        nc_amount_total.append(nc_total)
                        so_modified.append(inv_origin)
                        so_origin_invoice.append(inv_name)
                        so_mkp_reference.append(sale_ref)
                        progress_bar.update(1)
                else:
                    print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                    so_with_refund.append(inv_origin_name)
                    progress_bar.update(1)
                    continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_individuales_amazon_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['carlos.hinojosa@wonderbrands.co', 'sergio@wonderbrands.co', 'eric@wonderbrands.co', 'rosalba@wonderbrands.co', 'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Amazon - Creación de notas de crédito para facturas individuales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('PROCESO NC PARA AMAZON COMPLETADO :)')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
@current_execution
def reverse_invoice_global_amazon():
    # Formato para query
    type_filter = 'GLOBAL'
    marketplace_filter = 'AMAZON'
    list_orders, placeholders, num_records = e_o.filter_orders(orders_amz_file_path, type_filter, marketplace_filter)
    dates_list_params = [start_date_str, end_date_str]

    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('NOTAS DE CRÉDITO GLOBALES AMAZON')
    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('----------------------------------------------------------------')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')

    mycursor.execute("""#GLOBALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name/*,
                               d.order_id,
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               d.refunded_amt,
                               refund_date,
                               b.invoice_partner_display_name 'cliente',
                               'GLOBAL' as type,
                               'AMAZON' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN somos_reyes.odoo_new_sale_order c
                        ON SUBSTRING_INDEX(SUBSTRING_INDEX(invoice_ids, ']', 1), '[', -1) = b.id
                        LEFT JOIN (SELECT a.order_id, max(STR_TO_DATE(fecha, '%m/%d/%Y')) 'refund_date', SUM(total - tarifas_de_amazon) * (-1) 'refunded_amt'
                                   FROM somos_reyes.amazon_payments_refunds a
                                   WHERE (total - tarifas_de_amazon) * (-1) > 0 AND STR_TO_DATE(fecha, '%m/%d/%Y') >= %s AND STR_TO_DATE(fecha, '%m/%d/%Y') <= %s
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM somos_reyes.odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        WHERE d.order_id is not null
                        AND e.invoice_origin is null
                        AND invoice_partner_display_name = 'PÚBLICO EN GENERAL'
                        AND (d.refunded_amt - b.amount_total > 1 OR d.refunded_amt - b.amount_total < -1)
                        AND d.refunded_amt - c.amount_total < 1 AND d.refunded_amt - c.amount_total > -1
                        AND c.name in ({});
                        """.format(placeholders), tuple(dates_list_params+list_orders))
    invoice_records = mycursor.fetchall()
    #Lista de SO a las que se les creó una credit_notes
    so_modified = []
    #Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    #Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    #Lista de nombres de las notas de crédito creadas
    nc_created = []
    #Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    #Lista de facturas origen
    so_origin_invoice = []
    #Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    #Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0] # Almacena el nombre de la SO
            inv_id = each[1] # Almacena el ID de la factura
            inv_name = each[2] # Almacena el nombre de la factura
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_usage = inv['l10n_mx_edi_usage']  # Folio fiscal de la factura
                    inv_uuid_origin = f'03|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        #--------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        #Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund'], ['state', 'not ilike', 'cancel']]])
                        if not existing_credit_note:
                            #Busca la órden de venta
                            sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read', [[['name', '=', inv_origin_name]]])[0]
                            # Obtiene los datos necesarios directo de la SO
                            sale_id = sale_order['id']
                            sale_name = sale_order['name']
                            sale_ref = sale_order['channel_order_reference']
                            sale_team = sale_order['team_id'][0]
                            #Busca el order line correspondiente de la orden de venta
                            sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line', 'search_read', [[['order_id', '=', sale_id]]])
                            #Define los valores de la nota de crédito
                            inv_int = int(inv_id)
                            sale_int = int(sale_id)
                            refund_vals = {
                                'ref': f'Reversión de: {inv_name}',
                                'journal_id': inv_journal_id,
                                'invoice_origin': sale_name,
                                'team_id': sale_team,
                                'payment_reference': inv_name,
                                'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                # Puedes ajustar la fecha según tus necesidades
                                'partner_id': inv['partner_id'][0],
                                'l10n_mx_edi_usage': l10n_mx_edi_usage,
                                'l10n_mx_edi_origin': inv_uuid_origin,
                                'l10n_mx_edi_payment_method_id': l10n_mx_edi_payment_method_id,
                                'reversed_entry_id': inv_int,
                                'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                'invoice_line_ids': []
                            }
                            for lines in sale_line_id:
                                nc_lines = {'product_id': lines['product_id'][0],
                                            'quantity': lines['product_uom_qty'],
                                            'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                            'price_unit': lines['price_unit'],
                                            'product_uom_id': lines['product_uom'][0],
                                            'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                            }
                                refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                            #Crea la nota de crédito
                            create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create', [refund_vals])

                            #Agrega mensaje al Attachment de la nota de crédito
                            message = {
                                'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                'message_type': 'comment',
                            }
                            write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[create_nc], message)
                            #Enlazamos la venta con la nueva factura
                            upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write', [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                            #Publicamos la nota de crédito
                            upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                            #Timbramos la nota de crédito
                            # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                            stamp_credit_note(models, db_name, uid, password, create_nc)

                            #Buscamos el nombre de la factura ya creada
                            search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', create_nc]]])
                            nc_name = search_nc_name[0]['name']
                            nc_total = search_nc_name[0]['amount_total']
                            #Agregamos a las listas
                            so_modified.append(sale_name)
                            nc_created.append(nc_name)
                            nc_amount_total.append(nc_total)
                            so_origin_invoice.append(inv_name)
                            so_mkp_reference.append(sale_ref)
                            progress_bar.update(1)
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
       print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    #Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'notas_credito_globales_amazon_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'sergio@wonderbrands.co'
        msg['To'] = ', '.join(
            ['carlos.hinojosa@wonderbrands.co', 'sergio@wonderbrands.co', 'eric@wonderbrands.co', 'rosalba@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático Amazon - Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        #smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
        smtpObj.send_message(msg)
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC Amazon completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()


def stamp_credit_note(models,db_name,uid,password, credit_note_id):
    # Timbrar la Nota de credito (certificar)
    # invoice_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'action_l10n_mx_edi_invoice', [[invoice_id]])
    try:
        print('----------------------------------------------------------------')
        print('Timbrando nota de credito')
        credit_note_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'action_process_edi_web_services',
                                              [[credit_note_id]])
        print(f"Nota de crédito timbrada: {credit_note_stamp}")
    except Exception as e:
        print('----------------------------------------------------------------')
        print("Nota de credito timbrada")
    print('----------------------------------------------------------------')



if __name__ == "__main__":
    # Numero de workers = numero de funciones (para este script)
    num_workers = 2

    # Crear un ThreadPoolExecutor con `num_workers` hilos
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        # Enviar las funciones al executor
        futures = [
            #executor.submit(reverse_invoice_meli),
            #executor.submit(reverse_invoice_global_meli),
            executor.submit(reverse_invoice_amazon),
            executor.submit(reverse_invoice_global_amazon)
        ]

        # Esperar a que todas las funciones terminen
        for future in futures:
            future.result()

    end_time = datetime.datetime.now()
    duration = end_time - today_date
    print(f'Duraciòn del script: {duration}')
    print('Listo')
    print('Este arroz ya se coció :)')