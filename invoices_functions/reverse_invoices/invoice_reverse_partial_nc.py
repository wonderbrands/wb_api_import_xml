from email.message import EmailMessage
from email.utils import make_msgid
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from pprint import pprint
from email import encoders
from tqdm import tqdm
import time
import json
import jsonrpc
import jsonrpclib
import random
import urllib.request
import getpass
import http
import requests
import logging
import zipfile
import socket
import os
import locale
import xmlrpc.client
import base64
import openpyxl
import xlrd
import pandas as pd
import MySQLdb
import mysql.connector
import smtplib
import ssl
import email
import datetime

print('================================================================')
print('BIENVENIDO AL PROCESO DE NOTAS DE CRÉDITO PARA MARKETPLACES')
print('================================================================')
print('SCRIPT DE CREACIÓN DE NOTAS DE CRÉDITO PARA FACTURAS PARCIALES')
print('================================================================')
today_date = datetime.datetime.now()
dir_path = os.path.dirname(os.path.realpath(__file__))
print('Fecha:' + today_date.strftime("%Y-%m-%d %H:%M:%S"))
#Archivo de configuración - Use config_dev.json si está haciendo pruebas
#Archivo de configuración - Use config.json cuando los cambios vayan a producción
config_file_name = r'C:\Dev\wb_odoo_external_api\config\config.json'

def get_odoo_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['odoo']
def get_psql_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['psql']
def get_email_access():
    with open(config_file_name, 'r') as config_file:
        config = json.load(config_file)

    return config['email']
def reverse_invoice_partial_ind_meli():
    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    #GLOBALES MELI
    mycursor.execute("""#INDIVIDUALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name,
                               f.product_id,
                               ifnull(d.refunded_amt, dd.refunded_amt) 'ml_refunded_amount',
                               ROUND(ifnull(d.refunded_amt, dd.refunded_amt) / unit_price, 2) 'qty_refunded'/*,
                               ifnull(d.order_id, dd.pack_id) 'order_id_or_pack_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               ifnull(d.shipping_amt, dd.shipping_amt) 'ml_shipping_amount',
                               ifnull(d.payment_date_last_modified, dd.payment_date_last_modified) 'payment_date_last_modified',
                               b.invoice_partner_display_name 'cliente',
                               'INDIVIDUAL' as type,
                               'MERCADO LIBRE' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        
                        LEFT JOIN odoo_new_sale_order c
                        ON b.invoice_origin = c.name
                        
                        LEFT JOIN (SELECT a.order_id, sku_id,
                                          max(payment_date_last_modified) 'payment_date_last_modified',
                                          SUM(paid_amt) 'paid_amt',
                                          SUM(refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt',
                                          SUM(refunded_amt) / SUM(sku_unit_price) 'division',
                                          ROUND(SUM(refunded_amt) / SUM(sku_unit_price)) 'redondeo'
                                   FROM ml_order_payments a
                                   LEFT JOIN ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                                   GROUP BY 1, 2
                                   ) d
                        ON c.channel_order_id = d.order_id
                        
                        LEFT JOIN (SELECT a.pack_id, sku_id,
                                          max(payment_date_last_modified) 'payment_date_last_modified',
                                          SUM(b.paid_amt) 'paid_amt',
                                          SUM(b.refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt',
                                          SUM(refunded_amt) / SUM(sku_unit_price) 'division',
                                          ROUND(SUM(refunded_amt) / SUM(sku_unit_price)) 'redondeo'
                        FROM ml_order_update a
                        LEFT JOIN ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                        GROUP BY 1, 2
                        ) dd
                        ON c.yuju_pack_id = dd.pack_id
                        
                        LEFT JOIN (SELECT distinct invoice_origin FROM odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        
                        LEFT JOIN (SELECT order_name, default_code, product_id, SUM(product_qty) 'product_qty', ROUND(SUM(price_total) / SUM(product_qty), 2) 'unit_price'
                                   FROM odoo_new_sale_order_line a
                                   LEFT JOIN odoo_new_product_product_bis b
                                   ON a.product_id = b.id
                                   WHERE product_id <> '1'
                                   GROUP BY 1, 2, 3) f
                        ON c.name = f.order_name AND ifnull(d.sku_id, dd.sku_id) = f.default_code
                        
                        LEFT JOIN (SELECT a.order_id,
                                          SUM(refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt'
                                   FROM ml_order_payments a
                                   LEFT JOIN ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                                   GROUP BY 1) t
                        ON c.channel_order_id = t.order_id
                        
                        LEFT JOIN (SELECT a.pack_id,
                                          SUM(b.refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt'
                        FROM ml_order_update a
                        LEFT JOIN ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                        GROUP BY 1
                        ) tt
                        ON c.yuju_pack_id = tt.pack_id
                        
                        WHERE (d.order_id is not null or dd.pack_id is not null) #QUE TENGA REEMBLSO
                        AND e.invoice_origin is null #QUE NO TENGA NOTA DE CREDITO
                        AND b.invoice_partner_display_name <> 'PÚBLICO EN GENERAL'
                        AND c.amount_total - ifnull(t.refunded_amt, tt.refunded_amt) > 1 #QUE EL MONTO DEL REEMBOLSO SEA MENOR AL MONTO DE LA VENTA
                        AND c.amount_total - ifnull(t.refunded_amt - t.shipping_amt, tt.refunded_amt - tt.shipping_amt) > 1 #QUE EL MONTO DEL REEMBOLSO SEA MENOR AL MONTO DE LA VENTA, CONSIDERANDO ENVIO
                        AND (b.amount_total - c.amount_total < 1 AND b.amount_total - c.amount_total > (-1)) #QUE SEA INNDIVIDUAL
                        AND f.order_name is not null #QUE LA SO TENGA UN SOLO SKU
                        AND ROUND(ifnull(d.refunded_amt, dd.refunded_amt) / unit_price, 2) in (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20)
                        AND c.name in ('SO2317319', 'SO2328274', 'SO2330619', 'SO2360564', 'SO2516544', 'SO2285466', 'SO2306941', 'SO2265350', 'SO2378321', 'SO2425183', 'SO2252669', 'SO2319546', 'SO2410738', 'SO2386580', 'SO2384168', 'SO2374722', 'SO2371769', 'SO2420272', 'SO2416980', 'SO2275496', 'SO2322018', 'SO2352264', 'SO2400074', 'SO2361772', 'SO2250817', 'SO2240266', 'SO2321703', 'SO2228768', 'SO2283584', 'SO2354936', 'SO2273762', 'SO2244042', 'SO2276091', 'SO2396842', 'SO2381975', 'SO2401760', 'SO2297117', 'SO2298543', 'SO2368575', 'SO2332046', 'SO2247572', 'SO2357195', 'SO2256799', 'SO2304426', 'SO2322895', 'SO2321614', 'SO2324206', 'SO2380527', 'SO2255082', 'SO2404803', 'SO2268487', 'SO2303268', 'SO2338261', 'SO2400160', 'SO2388560', 'SO2272803', 'SO2266180', 'SO2334671', 'SO2354927', 'SO2291776', 'SO2280610', 'SO2270020', 'SO2246286', 'SO2274104', 'SO2377516', 'SO2148841', 'SO2360663', 'SO2284357', 'SO2258435', 'SO2349258', 'SO2331465', 'SO2402148', 'SO2230830', 'SO2252645', 'SO2342447', 'SO2260182', 'SO2451577', 'SO2419010', 'SO2260437', 'SO2314380', 'SO2365820', 'SO2370508', 'SO2343976', 'SO2247360', 'SO2328733', 'SO2289239', 'SO2319887', 'SO2354292', 'SO2380205', 'SO2258582', 'SO2270835', 'SO2312234', 'SO2312389', 'SO2284056', 'SO2324933', 'SO2230548', 'SO2407508', 'SO2304224', 'SO2281348', 'SO2132563', 'SO2178641', 'SO2324568', 'SO2295274', 'SO2267904', 'SO2321239', 'SO2264442', 'SO2377686', 'SO2337052', 'SO2356035', 'SO2301801', 'SO2313220', 'SO2394597', 'SO2245122', 'SO2244107', 'SO2257650', 'SO2318703', 'SO2356561', 'SO2309148', 'SO2305610', 'SO2428793', 'SO2245274', 'SO2350229', 'SO2149828', 'SO2302051', 'SO2332238', 'SO2365933', 'SO2364801', 'SO2247695', 'SO2394691', 'SO2345397', 'SO2296474', 'SO2416136', 'SO2417729', 'SO2279755', 'SO2274542', 'SO2313379', 'SO2233094', 'SO2446992', 'SO2434609', 'SO2424809', 'SO2438505', 'SO2438405', 'SO2437320', 'SO2428036', 'SO2443295', 'SO2443843', 'SO2425888', 'SO2425615', 'SO2446348', 'SO2421905', 'SO2437713', 'SO2480326', 'SO2233600', 'SO2252010', 'SO2456543', 'SO2481234', 'SO2452407', 'SO2436285', 'SO2459546', 'SO2429921', 'SO2420630', 'SO2460692', 'SO2466574', 'SO2467344', 'SO2466825', 'SO2462355', 'SO2464433', 'SO2466464', 'SO2453233', 'SO2462730', 'SO2467489', 'SO2484942', 'SO2468057', 'SO2471612', 'SO2472260', 'SO2459781', 'SO2449581', 'SO2494112', 'SO2483952', 'SO2483942', 'SO2483062', 'SO2484824', 'SO2457956', 'SO2480202', 'SO2500707', 'SO2501001', 'SO2497634', 'SO2547932', 'SO2542069', 'SO2529743', 'SO2584376', 'SO2518823', 'SO2556463', 'SO2553538', 'SO2553035', 'SO2495019', 'SO2495528', 'SO2515626', 'SO2526274', 'SO2513786', 'SO2512665', 'SO2510003', 'SO2515892', 'SO2517606', 'SO2517976', 'SO2527174', 'SO2520221', 'SO2539079', 'SO2513785', 'SO2469564', 'SO2529184', 'SO2474348', 'SO2526231', 'SO2520616', 'SO2484168', 'SO2500318', 'SO2527806', 'SO2494951', 'SO2524554', 'SO2532744', 'SO2521581', 'SO2525579', 'SO2531685', 'SO2535686')
                        """)
    invoice_records = mycursor.fetchall()
    #Lista de SO a las que se les creó una credit_notes
    so_modified = []
    #Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    #Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    #Lista de nombres de las notas de crédito creadas
    nc_created = []
    #Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    #Lista de facturas origen
    so_origin_invoice = []
    #Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    #Lista de SKUS dl reembolso
    nc_product_id = []
    #Lista de productos del reembolso que no existen en Odoo
    nc_product_id_no_exist = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    #Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0] # Almacena el nombre de la SO
            inv_id = each[1] # Almacena el ID de la factura
            inv_name = each[2] # Almacena el nombre de la factura
            inv_product_id = each[3] # Almacena el product_id del reembolso
            inv_refund_amount = float(each[4]) / 1.16 # Almacena el monto del reembolso
            inv_qty_refunded = each[5] # Almacena la cantidad del SKU reembolsado
            #Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_usage = 'G02'  # Uso del CFDI
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_uuid_origin = f'01|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    inv_payment = inv['l10n_mx_edi_payment_method_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        #--------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        #Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [[['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund']]])
                        if not existing_credit_note:
                            try:
                                #Busca la órden de venta
                                sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read', [[['name', '=', inv_origin_name]]])[0]
                                # Obtiene los datos necesarios directo de la SO
                                sale_id = sale_order['id']
                                sale_name = sale_order['name']
                                sale_ref = sale_order['channel_order_reference']
                                sale_team = sale_order['team_id'][0]
                                #Busca el order line correspondiente de la orden de venta
                                sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line', 'search_read', [[['order_id', '=', sale_id]]])
                                #Define los valores de la nota de crédito
                                inv_int = int(inv_id)
                                sale_int = int(sale_id)
                                refund_vals = {
                                    'ref': f'Reversión de: {inv_name}',
                                    'journal_id': inv_journal_id,
                                    'team_id': sale_team,
                                    'invoice_origin': sale_name,
                                    'payment_reference': inv_name,
                                    'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                    # Puedes ajustar la fecha según tus necesidades
                                    'partner_id': inv['partner_id'][0],
                                    'l10n_mx_edi_usage': inv_usage,
                                    'l10n_mx_edi_origin': inv_uuid_origin,
                                    'l10n_mx_edi_payment_method_id': inv_payment,
                                    'reversed_entry_id': inv_int,
                                    'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                    'invoice_line_ids': []
                                }
                                for lines in sale_line_id:
                                    if lines['product_id'][0] == int(inv_product_id):
                                        nc_lines = {'product_id': lines['product_id'][0],
                                                    'quantity': inv_qty_refunded,
                                                    'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                                    'price_unit': inv_refund_amount,
                                                    'product_uom_id': lines['product_uom'][0],
                                                    'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                                    }
                                        refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                                    else:
                                        print(f"El producto {inv_product_id} no coincide con ninguna línea de la factura {inv_name}")
                                        continue
                                #Crea la nota de crédito
                                create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create', [refund_vals])
                                #Actualiza la nota de crédito
                                #Agrega mensaje al Attachment de la nota de crédito
                                message = {
                                    'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                    'message_type': 'comment',
                                }
                                write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',[create_nc], message)
                                #Enlazamos la venta con la nueva factura
                                upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write', [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                                #Publicamos la nota de crédito
                                #upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                                #Timbramos la nota de crédito
                                #upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                                #Buscamos el nombre de la factura ya creada
                                search_nc_name = models.execute_kw(db_name, uid, password, 'account.move', 'search_read',[[['id', '=', create_nc]]])
                                nc_name = search_nc_name[0]['name']
                                nc_total = search_nc_name[0]['amount_total']
                                #Agregamos a las listas
                                so_modified.append(sale_name)
                                nc_created.append(nc_name)
                                nc_amount_total.append(nc_total)
                                so_origin_invoice.append(inv_name)
                                so_mkp_reference.append(sale_ref)
                                progress_bar.update(1)
                            except Exception as b:
                                print(f"En el armado de la factura y la creación: {b}")
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
       print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    #Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'nc_parciales_ind_meli_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales 
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito 
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron 
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'Tech anibal@wonderbrands.co'
        msg['To'] = ', '.join(
            ['anibal@wonderbrands.co', 'rosalba@wonderbrands.co', 'natalia@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático MELI- Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC globales Meli completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
def reverse_invoice_partial_glob_meli():
    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    # GLOBALES MELI
    mycursor.execute("""#GLOBALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name,
                               f.product_id,
                               ifnull(d.refunded_amt, dd.refunded_amt) 'ml_refunded_amount',
                               ROUND(ifnull(d.refunded_amt, dd.refunded_amt) / unit_price, 2) 'qty_refunded'/*,
                               ifnull(d.order_id, dd.pack_id) 'order_id_or_pack_id',
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               c.amount_total 'total_so',
                               ifnull(d.payment_date_last_modified, dd.payment_date_last_modified) 'payment_date_last_modified',
                               b.invoice_partner_display_name 'cliente',
                               'GLOBAL' as type,
                               'MERCADO LIBRE' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN odoo_new_sale_order c
                        ON SUBSTRING_INDEX(SUBSTRING_INDEX(invoice_ids, ']', 1), '[', -1) = b.id
                        LEFT JOIN (SELECT a.order_id, sku_id,
                                          max(payment_date_last_modified) 'payment_date_last_modified',
                                          SUM(paid_amt) 'paid_amt',
                                          SUM(refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt',
                                          SUM(refunded_amt) / SUM(sku_unit_price) 'division',
                                          ROUND(SUM(refunded_amt) / SUM(sku_unit_price)) 'redondeo'
                                   FROM ml_order_payments a
                                   LEFT JOIN ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                                   GROUP BY 1, 2
                                   ) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT a.pack_id, sku_id,
                                          max(payment_date_last_modified) 'payment_date_last_modified',
                                          SUM(b.paid_amt) 'paid_amt',
                                          SUM(b.refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt',
                                          SUM(refunded_amt) / SUM(sku_unit_price) 'division',
                                          ROUND(SUM(refunded_amt) / SUM(sku_unit_price)) 'redondeo'
                        FROM ml_order_update a
                        LEFT JOIN ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                        GROUP BY 1, 2
                        ) dd
                        ON c.yuju_pack_id = dd.pack_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        LEFT JOIN (SELECT order_name, default_code, product_id, SUM(product_qty) 'product_qty', ROUND(SUM(price_total) / SUM(product_qty), 2) 'unit_price'
                                   FROM odoo_new_sale_order_line a
                                   LEFT JOIN odoo_new_product_product_bis b
                                   ON a.product_id = b.id
                                   WHERE product_id <> '1'
                                   GROUP BY 1, 2, 3) f
                        ON c.name = f.order_name AND ifnull(d.sku_id, dd.sku_id) = f.default_code
                        LEFT JOIN (SELECT a.order_id,
                                          SUM(refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt'
                                   FROM ml_order_payments a
                                   LEFT JOIN ml_order_update b
                                   ON a.order_id = b.order_id
                                   WHERE refunded_amt > 0 AND b.pack_id = 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                                   GROUP BY 1) t
                        ON c.channel_order_id = t.order_id
                        
                        LEFT JOIN (SELECT a.pack_id,
                                          SUM(b.refunded_amt) 'refunded_amt',
                                          SUM(shipping_amt) 'shipping_amt'
                        FROM ml_order_update a
                        LEFT JOIN ml_order_payments b
                        ON a.order_id = b.order_id
                        WHERE b.refunded_amt > 0 AND a.pack_id <> 'None' AND date(payment_date_last_modified) >= '2023-01-01' AND date(payment_date_last_modified) <= '2023-11-30'
                        GROUP BY 1
                        ) tt
                        ON c.yuju_pack_id = tt.pack_id
                        WHERE (d.order_id is not null or dd.pack_id is not null)
                        AND e.invoice_origin is null
                        AND invoice_partner_display_name = 'PÚBLICO EN GENERAL'
                        AND c.amount_total - ifnull(t.refunded_amt, tt.refunded_amt) > 1
                        AND c.amount_total - ifnull(t.refunded_amt - t.shipping_amt, tt.refunded_amt - tt.shipping_amt) > 1
                        AND (b.amount_total - c.amount_total > 1 OR b.amount_total - c.amount_total < (-1))
                        AND f.order_name is not null
                        AND ROUND(ifnull(d.refunded_amt, dd.refunded_amt) / unit_price, 2) in (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20)
                        AND c.name in ('SO2317319', 'SO2328274', 'SO2330619', 'SO2360564', 'SO2516544', 'SO2285466', 'SO2306941', 'SO2265350', 'SO2378321', 'SO2425183', 'SO2252669', 'SO2319546', 'SO2410738', 'SO2386580', 'SO2384168', 'SO2374722', 'SO2371769', 'SO2420272', 'SO2416980', 'SO2275496', 'SO2322018', 'SO2352264', 'SO2400074', 'SO2361772', 'SO2250817', 'SO2240266', 'SO2321703', 'SO2228768', 'SO2283584', 'SO2354936', 'SO2273762', 'SO2244042', 'SO2276091', 'SO2396842', 'SO2381975', 'SO2401760', 'SO2297117', 'SO2298543', 'SO2368575', 'SO2332046', 'SO2247572', 'SO2357195', 'SO2256799', 'SO2304426', 'SO2322895', 'SO2321614', 'SO2324206', 'SO2380527', 'SO2255082', 'SO2404803', 'SO2268487', 'SO2303268', 'SO2338261', 'SO2400160', 'SO2388560', 'SO2272803', 'SO2266180', 'SO2334671', 'SO2354927', 'SO2291776', 'SO2280610', 'SO2270020', 'SO2246286', 'SO2274104', 'SO2377516', 'SO2148841', 'SO2360663', 'SO2284357', 'SO2258435', 'SO2349258', 'SO2331465', 'SO2402148', 'SO2230830', 'SO2252645', 'SO2342447', 'SO2260182', 'SO2451577', 'SO2419010', 'SO2260437', 'SO2314380', 'SO2365820', 'SO2370508', 'SO2343976', 'SO2247360', 'SO2328733', 'SO2289239', 'SO2319887', 'SO2354292', 'SO2380205', 'SO2258582', 'SO2270835', 'SO2312234', 'SO2312389', 'SO2284056', 'SO2324933', 'SO2230548', 'SO2407508', 'SO2304224', 'SO2281348', 'SO2132563', 'SO2178641', 'SO2324568', 'SO2295274', 'SO2267904', 'SO2321239', 'SO2264442', 'SO2377686', 'SO2337052', 'SO2356035', 'SO2301801', 'SO2313220', 'SO2394597', 'SO2245122', 'SO2244107', 'SO2257650', 'SO2318703', 'SO2356561', 'SO2309148', 'SO2305610', 'SO2428793', 'SO2245274', 'SO2350229', 'SO2149828', 'SO2302051', 'SO2332238', 'SO2365933', 'SO2364801', 'SO2247695', 'SO2394691', 'SO2345397', 'SO2296474', 'SO2416136', 'SO2417729', 'SO2279755', 'SO2274542', 'SO2313379', 'SO2233094', 'SO2446992', 'SO2434609', 'SO2424809', 'SO2438505', 'SO2438405', 'SO2437320', 'SO2428036', 'SO2443295', 'SO2443843', 'SO2425888', 'SO2425615', 'SO2446348', 'SO2421905', 'SO2437713', 'SO2480326', 'SO2233600', 'SO2252010', 'SO2456543', 'SO2481234', 'SO2452407', 'SO2436285', 'SO2459546', 'SO2429921', 'SO2420630', 'SO2460692', 'SO2466574', 'SO2467344', 'SO2466825', 'SO2462355', 'SO2464433', 'SO2466464', 'SO2453233', 'SO2462730', 'SO2467489', 'SO2484942', 'SO2468057', 'SO2471612', 'SO2472260', 'SO2459781', 'SO2449581', 'SO2494112', 'SO2483952', 'SO2483942', 'SO2483062', 'SO2484824', 'SO2457956', 'SO2480202', 'SO2500707', 'SO2501001', 'SO2497634', 'SO2547932', 'SO2542069', 'SO2529743', 'SO2584376', 'SO2518823', 'SO2556463', 'SO2553538', 'SO2553035', 'SO2495019', 'SO2495528', 'SO2515626', 'SO2526274', 'SO2513786', 'SO2512665', 'SO2510003', 'SO2515892', 'SO2517606', 'SO2517976', 'SO2527174', 'SO2520221', 'SO2539079', 'SO2513785', 'SO2469564', 'SO2529184', 'SO2474348', 'SO2526231', 'SO2520616', 'SO2484168', 'SO2500318', 'SO2527806', 'SO2494951', 'SO2524554', 'SO2532744', 'SO2521581', 'SO2525579', 'SO2531685', 'SO2535686')
                        """)
    invoice_records = mycursor.fetchall()
    # Lista de SO a las que se les creó una credit_notes
    so_modified = []
    # Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    # Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    # Lista de nombres de las notas de crédito creadas
    nc_created = []
    # Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    # Lista de facturas origen
    so_origin_invoice = []
    # Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    # Lista de SKUS dl reembolso
    nc_product_id = []
    # Lista de productos del reembolso que no existen en Odoo
    nc_product_id_no_exist = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    # Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0]  # Almacena el nombre de la SO
            inv_id = each[1]  # Almacena el ID de la factura
            inv_name = each[2]  # Almacena el nombre de la factura
            inv_product_id = each[3]  # Almacena el product_id del reembolso
            inv_refund_amount = float(each[4]) / 1.16  # Almacena el monto del reembolso
            inv_qty_refunded = each[5]  # Almacena la cantidad del SKU reembolsado
            # Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_usage = 'G02'  # Uso del CFDI
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_uuid_origin = f'01|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    inv_payment = inv['l10n_mx_edi_payment_method_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        # --------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        # Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [
                            [['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund']]])
                        if not existing_credit_note:
                            try:
                                # Busca la órden de venta
                                sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read',
                                                               [[['name', '=', inv_origin_name]]])[0]
                                # Obtiene los datos necesarios directo de la SO
                                sale_id = sale_order['id']
                                sale_name = sale_order['name']
                                sale_ref = sale_order['channel_order_reference']
                                sale_team = sale_order['team_id'][0]
                                # Busca el order line correspondiente de la orden de venta
                                sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line','search_read', [[['order_id', '=', sale_id]]])
                                # Define los valores de la nota de crédito
                                inv_int = int(inv_id)
                                sale_int = int(sale_id)
                                refund_vals = {
                                    'ref': f'Reversión de: {inv_name}',
                                    'journal_id': inv_journal_id,
                                    'team_id': sale_team,
                                    'invoice_origin': sale_name,
                                    'payment_reference': inv_name,
                                    'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                    # Puedes ajustar la fecha según tus necesidades
                                    'partner_id': inv['partner_id'][0],
                                    'l10n_mx_edi_usage': inv_usage,
                                    'l10n_mx_edi_origin': inv_uuid_origin,
                                    'l10n_mx_edi_payment_method_id': inv_payment,
                                    'reversed_entry_id': inv_int,
                                    'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                    'invoice_line_ids': []
                                }
                                for lines in sale_line_id:
                                    if lines['product_id'][0] == int(inv_product_id):
                                        nc_lines = {'product_id': lines['product_id'][0],
                                                    'quantity': inv_qty_refunded,
                                                    'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                                    'price_unit': lines['price_unit'],
                                                    'product_uom_id': lines['product_uom'][0],
                                                    'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                                    }
                                        refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                                    else:
                                        #print(f"El producto {inv_product_id} no coincide con ninguna línea de la factura {inv_name}")
                                        continue
                                # Crea la nota de crédito
                                create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create',
                                                              [refund_vals])
                                # Actualiza la nota de crédito
                                # Agrega mensaje al Attachment de la nota de crédito
                                message = {
                                    'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                    'message_type': 'comment',
                                }
                                write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',
                                                                 [create_nc], message)
                                # Enlazamos la venta con la nueva factura
                                upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write',
                                                             [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                                # Publicamos la nota de crédito
                                # upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                                # Timbramos la nota de crédito
                                # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                                # Buscamos el nombre de la factura ya creada
                                search_nc_name = models.execute_kw(db_name, uid, password, 'account.move',
                                                                   'search_read', [[['id', '=', create_nc]]])
                                nc_name = search_nc_name[0]['name']
                                nc_total = search_nc_name[0]['amount_total']
                                # Agregamos a las listas
                                so_modified.append(sale_name)
                                nc_created.append(nc_name)
                                nc_amount_total.append(nc_total)
                                so_origin_invoice.append(inv_name)
                                so_mkp_reference.append(sale_ref)
                                progress_bar.update(1)
                            except Exception as b:
                                print(f"En el armado de la factura y la creación: {b}")
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    # Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'nc_parciales_glo_meli_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales 
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito 
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron 
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'Tech anibal@wonderbrands.co'
        msg['To'] = ', '.join(
            ['anibal@wonderbrands.co', 'rosalba@wonderbrands.co', 'natalia@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático MELI- Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC globales Meli completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
def reverse_invoice_partial_ind_amz():
    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    # GLOBALES MELI
    mycursor.execute("""#INDIVIDUALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name,
                               f.product_id,
                               d.refunded_amt,
                               ROUND(d.refunded_amt / unit_price, 2) 'qty_refunded'/*,
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               d.order_id 'order_id',
                               d.refund_date,
                               b.invoice_partner_display_name 'cliente',
                               'INDIVIDUAL' as type,
                               'AMAZON' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        
                        LEFT JOIN odoo_new_sale_order c
                        ON b.invoice_origin = c.name
                        
                        LEFT JOIN (SELECT a.order_id, max(STR_TO_DATE(fecha, '%d/%m/%Y')) 'refund_date', SUM(total - tarifas_de_amazon) * (-1) 'refunded_amt'
                                   FROM somos_reyes.amazon_payments_refunds a
                                   WHERE (total - tarifas_de_amazon) * (-1) > 0 AND STR_TO_DATE(fecha, '%d/%m/%Y') >= '2023-01-01' AND STR_TO_DATE(fecha, '%d/%m/%Y') <= '2023-11-30'
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        
                        LEFT JOIN (SELECT distinct invoice_origin FROM odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        
                        LEFT JOIN (SELECT order_name, MAX(product_id) 'product_id', SUM(product_qty) 'product_qty', COUNT(distinct product_id) 'cuenta', SUM(price_total) / SUM(product_qty) 'unit_price'
                                   FROM odoo_new_sale_order_line a
                                   LEFT JOIN odoo_new_product_product_bis b
                                   ON a.product_id = b.id
                                   WHERE product_id <> '1'
                                   GROUP BY 1
                                   HAVING cuenta = 1) f
                        ON c.name = f.order_name
                        
                        WHERE d.order_id is not null
                        AND e.invoice_origin is null
                        
                        AND c.amount_total - d.refunded_amt > 1 #QUE EL REEMBOLSO SEA MENOR A LA SO
                        AND (b.amount_total - c.amount_total < 1 AND b.amount_total - c.amount_total > (-1)) #QUE SEA INDIVIDUAL
                        AND f.order_name is not null
                        AND ROUND(d.refunded_amt / unit_price, 2) in (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20)
                        """)
    invoice_records = mycursor.fetchall()
    # Lista de SO a las que se les creó una credit_notes
    so_modified = []
    # Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    # Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    # Lista de nombres de las notas de crédito creadas
    nc_created = []
    # Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    # Lista de facturas origen
    so_origin_invoice = []
    # Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    # Lista de SKUS dl reembolso
    nc_product_id = []
    # Lista de productos del reembolso que no existen en Odoo
    nc_product_id_no_exist = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    # Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0]  # Almacena el nombre de la SO
            inv_id = each[1]  # Almacena el ID de la factura
            inv_name = each[2]  # Almacena el nombre de la factura
            inv_product_id = each[3]  # Almacena el product_id del reembolso
            inv_refund_amount = float(each[4]) / 1.16  # Almacena el monto del reembolso
            inv_qty_refunded = each[5]  # Almacena la cantidad del SKU reembolsado
            # Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_usage = 'G02'  # Uso del CFDI
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_uuid_origin = f'01|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    inv_payment = inv['l10n_mx_edi_payment_method_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        # --------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        # Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [
                            [['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund']]])
                        if not existing_credit_note:
                            try:
                                # Busca la órden de venta
                                sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read',
                                                               [[['name', '=', inv_origin_name]]])[0]
                                # Obtiene los datos necesarios directo de la SO
                                sale_id = sale_order['id']
                                sale_name = sale_order['name']
                                sale_ref = sale_order['channel_order_reference']
                                sale_team = sale_order['team_id'][0]
                                # Busca el order line correspondiente de la orden de venta
                                sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line',
                                                                 'search_read', [[['order_id', '=', sale_id]]])
                                # Define los valores de la nota de crédito
                                inv_int = int(inv_id)
                                sale_int = int(sale_id)
                                refund_vals = {
                                    'ref': f'Reversión de: {inv_name}',
                                    'journal_id': inv_journal_id,
                                    'team_id': sale_team,
                                    'invoice_origin': sale_name,
                                    'payment_reference': inv_name,
                                    'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                    # Puedes ajustar la fecha según tus necesidades
                                    'partner_id': inv['partner_id'][0],
                                    'l10n_mx_edi_usage': inv_usage,
                                    'l10n_mx_edi_origin': inv_uuid_origin,
                                    'l10n_mx_edi_payment_method_id': inv_payment,
                                    'reversed_entry_id': inv_int,
                                    'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                    'invoice_line_ids': []
                                }
                                for lines in sale_line_id:
                                    if lines['product_id'][0] == int(inv_product_id):
                                        nc_lines = {'product_id': lines['product_id'][0],
                                                    'quantity': inv_qty_refunded,
                                                    'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                                    'price_unit': lines['price_unit'],
                                                    'product_uom_id': lines['product_uom'][0],
                                                    'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                                    }
                                        refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                                    else:
                                        # print(f"El producto {inv_product_id} no coincide con ninguna línea de la factura {inv_name}")
                                        continue
                                # Crea la nota de crédito
                                create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create',
                                                              [refund_vals])
                                # Actualiza la nota de crédito
                                # Agrega mensaje al Attachment de la nota de crédito
                                message = {
                                    'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                    'message_type': 'comment',
                                }
                                write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',
                                                                 [create_nc], message)
                                # Enlazamos la venta con la nueva factura
                                upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write',
                                                             [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                                # Publicamos la nota de crédito
                                # upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                                # Timbramos la nota de crédito
                                # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                                # Buscamos el nombre de la factura ya creada
                                search_nc_name = models.execute_kw(db_name, uid, password, 'account.move',
                                                                   'search_read', [[['id', '=', create_nc]]])
                                nc_name = search_nc_name[0]['name']
                                nc_total = search_nc_name[0]['amount_total']
                                # Agregamos a las listas
                                so_modified.append(sale_name)
                                nc_created.append(nc_name)
                                nc_amount_total.append(nc_total)
                                so_origin_invoice.append(inv_name)
                                so_mkp_reference.append(sale_ref)
                                progress_bar.update(1)
                            except Exception as b:
                                print(f"En el armado de la factura y la creación: {b}")
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    # Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'nc_parciales_ind_amz_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales 
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito 
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron 
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'Tech anibal@wonderbrands.co'
        msg['To'] = ', '.join(
            ['anibal@wonderbrands.co', 'rosalba@wonderbrands.co', 'natalia@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático MELI- Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC globales Meli completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()
def reverse_invoice_partial_glo_amz():
    # Obtener credenciales
    odoo_keys = get_odoo_access()
    psql_keys = get_psql_access()
    email_keys = get_email_access()
    # odoo
    server_url = odoo_keys['odoourl']
    db_name = odoo_keys['odoodb']
    username = odoo_keys['odoouser']
    password = odoo_keys['odoopassword']
    # correo
    smtp_server = email_keys['smtp_server']
    smtp_port = email_keys['smtp_port']
    smtp_username = email_keys['smtp_username']
    smtp_password = email_keys['smtp_password']

    print('----------------------------------------------------------------')
    print('Conectando API Odoo')
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(server_url))
    uid = common.authenticate(db_name, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(server_url))
    print('Conexión con Odoo establecida')
    print('Conectando a Mysql')
    # Connect to MySQL database
    mydb = mysql.connector.connect(
        host=psql_keys['dbhost'],
        user=psql_keys['dbuser'],
        password=psql_keys['dbpassword'],
        database=psql_keys['database']
    )
    mycursor = mydb.cursor()
    print('----------------------------------------------------------------')
    print('Vaya por un tecito o un café porque este proceso tomará algo de tiempo')
    # GLOBALES MELI
    mycursor.execute("""#GLOBALES
                        SELECT c.name,
                               b.id 'account_move_id',
                               b.name,
                               f.product_id,
                               d.refunded_amt,
                               ROUND(d.refunded_amt / unit_price, 2) 'qty_refunded'/*,
                               d.order_id,
                               b.amount_total 'total_factura',
                               b.amount_untaxed 'subtotal_factura',
                               c.amount_total 'so_amount',
                               refund_date,
                               b.invoice_partner_display_name 'cliente',
                               'GLOBAL' as type,
                               'AMAZON' as marketplace*/
                        FROM somos_reyes.odoo_new_account_move_aux b
                        LEFT JOIN odoo_new_sale_order c
                        ON SUBSTRING_INDEX(SUBSTRING_INDEX(invoice_ids, ']', 1), '[', -1) = b.id
                        LEFT JOIN (SELECT a.order_id, max(STR_TO_DATE(fecha, '%d/%m/%Y')) 'refund_date', SUM(total - tarifas_de_amazon) * (-1) 'refunded_amt'
                                   FROM somos_reyes.amazon_payments_refunds a
                                   WHERE (total - tarifas_de_amazon) * (-1) > 0 AND STR_TO_DATE(fecha, '%d/%m/%Y') >= '2023-01-01' AND STR_TO_DATE(fecha, '%d/%m/%Y') <= '2023-11-30'
                                   GROUP BY 1) d
                        ON c.channel_order_id = d.order_id
                        LEFT JOIN (SELECT distinct invoice_origin FROM odoo_new_account_move_aux WHERE name like '%RINV%') e
                        ON c.name = e.invoice_origin
                        LEFT JOIN (SELECT order_name, MAX(product_id) 'product_id', SUM(product_qty) 'product_qty', COUNT(distinct product_id) 'cuenta', SUM(price_total) / SUM(product_qty) 'unit_price'
                                   FROM odoo_new_sale_order_line
                                   WHERE product_id <> '1'
                                   GROUP BY 1
                                   HAVING cuenta = 1) f
                        ON c.name = f.order_name
                        WHERE d.order_id is not null
                        AND e.invoice_origin is null
                        AND invoice_partner_display_name = 'PÚBLICO EN GENERAL'
                        AND c.amount_total - d.refunded_amt > 1
                        AND (b.amount_total - c.amount_total > 1 OR b.amount_total - c.amount_total < (-1))
                        AND f.order_name is not null
                        AND ROUND(d.refunded_amt / unit_price, 2) in (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20)
                        """)
    invoice_records = mycursor.fetchall()
    # Lista de SO a las que se les creó una credit_notes
    so_modified = []
    # Lista de las facturas enlazadas a la SO y no existen
    inv_no_exist = []
    # Lista de SO que ya contaban con credit_notes antes del script
    so_with_refund = []
    # Lista de nombres de las notas de crédito creadas
    nc_created = []
    # Lista de SO que no existen en la factura global que tienen enlazada
    so_no_exist_in_invoice = []
    # Lista de facturas origen
    so_origin_invoice = []
    # Lista de referencias MKP para cada SO
    so_mkp_reference = []
    # Lista de total de la NC
    nc_amount_total = []
    # Lista de SKUS dl reembolso
    nc_product_id = []
    # Lista de productos del reembolso que no existen en Odoo
    nc_product_id_no_exist = []
    print('----------------------------------------------------------------')
    print('Creando notas de crédito')
    print('Este proceso tomará unos minutos')
    # Creación de notas de crédito
    try:
        progress_bar = tqdm(total=len(invoice_records), desc="Procesando")
        for each in invoice_records:
            inv_origin_name = each[0]  # Almacena el nombre de la SO
            inv_id = each[1]  # Almacena el ID de la factura
            inv_name = each[2]  # Almacena el nombre de la factura
            inv_product_id = each[3]  # Almacena el product_id del reembolso
            inv_refund_amount = float(each[4]) / 1.16  # Almacena el monto del reembolso
            inv_qty_refunded = each[5]  # Almacena la cantidad del SKU reembolsado
            # Busca la factura que contenga el nombre de la SO
            invoice = models.execute_kw(db_name, uid, password, 'account.move', 'search_read', [[['id', '=', inv_id]]])
            if invoice:
                for inv in invoice:
                    inv_usage = 'G02'  # Uso del CFDI
                    inv_uuid = inv['l10n_mx_edi_cfdi_uuid']  # Folio fiscal de la factura
                    inv_uuid_origin = f'01|{inv_uuid}'
                    inv_journal_id = inv['journal_id'][0]
                    inv_payment = inv['l10n_mx_edi_payment_method_id'][0]
                    if inv_origin_name in inv['invoice_origin']:
                        # --------------------------AGREGAR CONDICIONAL PARA SABER SI TIENE NOTA DE CREDITO--------------------------
                        # Validamos si la SO ya tiene una nota de crédito creada
                        existing_credit_note = models.execute_kw(db_name, uid, password, 'account.move', 'search', [
                            [['invoice_origin', '=', inv_origin_name], ['move_type', '=', 'out_refund']]])
                        if not existing_credit_note:
                            try:
                                # Busca la órden de venta
                                sale_order = models.execute_kw(db_name, uid, password, 'sale.order', 'search_read',
                                                               [[['name', '=', inv_origin_name]]])[0]
                                # Obtiene los datos necesarios directo de la SO
                                sale_id = sale_order['id']
                                sale_name = sale_order['name']
                                sale_ref = sale_order['channel_order_reference']
                                sale_team = sale_order['team_id'][0]
                                # Busca el order line correspondiente de la orden de venta
                                sale_line_id = models.execute_kw(db_name, uid, password, 'sale.order.line',
                                                                 'search_read', [[['order_id', '=', sale_id]]])
                                # Define los valores de la nota de crédito
                                inv_int = int(inv_id)
                                sale_int = int(sale_id)
                                refund_vals = {
                                    'ref': f'Reversión de: {inv_name}',
                                    'journal_id': inv_journal_id,
                                    'team_id': sale_team,
                                    'invoice_origin': sale_name,
                                    'payment_reference': inv_name,
                                    'invoice_date': datetime.datetime.now().strftime('%Y-%m-%d'),
                                    # Puedes ajustar la fecha según tus necesidades
                                    'partner_id': inv['partner_id'][0],
                                    'l10n_mx_edi_usage': inv_usage,
                                    'l10n_mx_edi_origin': inv_uuid_origin,
                                    'l10n_mx_edi_payment_method_id': inv_payment,
                                    'reversed_entry_id': inv_int,
                                    'move_type': 'out_refund',  # Este campo indica que es una nota de crédito
                                    'invoice_line_ids': []
                                }
                                for lines in sale_line_id:
                                    if lines['product_id'][0] == int(inv_product_id):
                                        nc_lines = {'product_id': lines['product_id'][0],
                                                    'quantity': inv_qty_refunded,
                                                    'name': lines['name'],  # Puedes ajustar esto según tus necesidades
                                                    'price_unit': lines['price_unit'],
                                                    'product_uom_id': lines['product_uom'][0],
                                                    'tax_ids': [(6, 0, [lines['tax_id'][0]])],
                                                    }
                                        refund_vals['invoice_line_ids'].append((0, 0, nc_lines))
                                    else:
                                        # print(f"El producto {inv_product_id} no coincide con ninguna línea de la factura {inv_name}")
                                        continue
                                # Crea la nota de crédito
                                create_nc = models.execute_kw(db_name, uid, password, 'account.move', 'create',
                                                              [refund_vals])
                                # Actualiza la nota de crédito
                                # Agrega mensaje al Attachment de la nota de crédito
                                message = {
                                    'body': f"Esta nota de crédito fue creada a partir de la factura: {inv_name}, de la órden {sale_name}, con folio fiscal {inv_uuid}, a solicitud del equipo de Contabilidad, por el equipo de Tech mediante API.",
                                    'message_type': 'comment',
                                }
                                write_msg_nc = models.execute_kw(db_name, uid, password, 'account.move', 'message_post',
                                                                 [create_nc], message)
                                # Enlazamos la venta con la nueva factura
                                upd_sale = models.execute_kw(db_name, uid, password, 'sale.order', 'write',
                                                             [[sale_id], {'invoice_ids': [(4, 0, create_nc)]}])
                                # Publicamos la nota de crédito
                                # upd_nc_state = models.execute_kw(db_name, uid, password, 'account.move', 'action_post', [create_nc])
                                # Timbramos la nota de crédito
                                # upd_nc_stamp = models.execute_kw(db_name, uid, password, 'account.move', 'button_process_edi_web_services',[create_nc])
                                # Buscamos el nombre de la factura ya creada
                                search_nc_name = models.execute_kw(db_name, uid, password, 'account.move',
                                                                   'search_read', [[['id', '=', create_nc]]])
                                nc_name = search_nc_name[0]['name']
                                nc_total = search_nc_name[0]['amount_total']
                                # Agregamos a las listas
                                so_modified.append(sale_name)
                                nc_created.append(nc_name)
                                nc_amount_total.append(nc_total)
                                so_origin_invoice.append(inv_name)
                                so_mkp_reference.append(sale_ref)
                                progress_bar.update(1)
                            except Exception as b:
                                print(f"En el armado de la factura y la creación: {b}")
                        else:
                            print(f"La órden {inv_origin_name} ya tiene una nota de crédito creada")
                            so_with_refund.append(inv_origin_name)
                            progress_bar.update(1)
                            continue
                    else:
                        print(f"La órden {inv_origin_name} no se encontró en la factura global")
                        so_no_exist_in_invoice.append(inv_origin_name)
                        progress_bar.update(1)
                        continue
            else:
                print(f"No hay una factura en la SO {inv_origin_name} por la cual se pueda crear una nota de crédito")
                inv_no_exist.append(inv_origin_name)
                progress_bar.update(1)
                continue
    except Exception as e:
        print(f"Error: no se pudo crear la nota de crédito: {e}")
    # Define el cuerpo del correo
    print('----------------------------------------------------------------')
    print('Creando correo y excel')
    # Excel
    try:
        # Crear el archivo Excel y agregar los nombres de los arrays y los resultados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'so_modified'
        sheet['B1'] = 'so_mkp_reference'
        sheet['C1'] = 'nc_created'
        sheet['D1'] = 'nc_amount_total'
        sheet['E1'] = 'so_origin_invoice'
        sheet['F1'] = 'inv_no_exist'
        sheet['G1'] = 'so_with_refund'
        sheet['H1'] = 'so_no_exist_in_invoice'

        # Agregar los resultados de los arrays
        for i in range(len(so_modified)):
            sheet['A{}'.format(i + 2)] = so_modified[i]
        for i in range(len(so_mkp_reference)):
            sheet['B{}'.format(i + 2)] = so_mkp_reference[i]
        for i in range(len(nc_created)):
            sheet['C{}'.format(i + 2)] = nc_created[i]
        for i in range(len(nc_amount_total)):
            sheet['D{}'.format(i + 2)] = nc_amount_total[i]
        for i in range(len(so_origin_invoice)):
            sheet['E{}'.format(i + 2)] = so_origin_invoice[i]
        for i in range(len(inv_no_exist)):
            sheet['F{}'.format(i + 2)] = inv_no_exist[i]
        for i in range(len(so_with_refund)):
            sheet['G{}'.format(i + 2)] = so_with_refund[i]
        for i in range(len(so_no_exist_in_invoice)):
            sheet['H{}'.format(i + 2)] = so_no_exist_in_invoice[i]

        # Guardar el archivo Excel en disco
        excel_file = 'nc_parciales_glo_amz_' + today_date.strftime("%Y%m%d") + '.xlsx'
        workbook.save(excel_file)

        # Leer el contenido del archivo Excel
        with open(excel_file, 'rb') as file:
            file_data = file.read()
        file_data_encoded = base64.b64encode(file_data).decode('utf-8')
    except Exception as a:
        print(f"Error: no se pudo crear el archivo de excel: {a}")
    # Correo
    try:
        msg = MIMEMultipart()
        body = '''\
                <html>
                  <head></head>
                  <body>
                    <p>Buenas</p>
                    <p>Hola a todos, espero que estén muy bien. Les comento que acabamos de correr el script de notas de crédito.</p>
                    <p>Adjunto encontrarán el archivo generado por el script en el cual se encuentran las órdenes a las cuales 
                    se les creó una nota de crédito, órdenes que no se les pudo crear una credit_notes, nombre de las notas de crédito 
                    creadas, órdenes que ya contaban con una nota de crédito antes de correr el script y órdenes que tuvieron 
                    algún error, por ejemplo que no existieran dentro de la factura global o no tuvieran una factura creada por la cual se pueda emitir una nota de crédito.</p>
                    </br>
                    <p>Sin más por el momento quedo al pendiente para resolver cualquier duda o comentario.</p>
                    </br>
                    <p>Muchas gracias</p>
                    </br>
                    <p>Un abrazo</p>
                  </body>
                </html>
                '''
        # Define remitente y destinatario
        msg = MIMEMultipart()
        msg['From'] = 'Tech anibal@wonderbrands.co'
        msg['To'] = ', '.join(
            ['anibal@wonderbrands.co', 'rosalba@wonderbrands.co', 'natalia@wonderbrands.co',
             'greta@somos-reyes.com',
             'contabilidad@somos-reyes.com', 'alex@wonderbrands.co', 'will@wonderbrands.co'])
        msg['Subject'] = 'Script Automático MELI- Creación de notas de crédito para facturas globales'
        # Adjuntar el cuerpo del correo
        msg.attach(MIMEText(body, 'html'))
        # Adjuntar el archivo Excel al mensaje
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attachment)
        print("Enviando correo")
        smtpObj = smtplib.SMTP(smtp_server, smtp_port)
        smtpObj.starttls()
        smtpObj.login(smtp_username, smtp_password)
        smtpObj.sendmail(smtp_username, msg['To'], msg.as_string())
    except Exception as i:
        print(f"Error: no se pudo enviar el correo: {i}")

    print('----------------------------------------------------------------')
    print('Proceso NC globales Meli completado')
    print('----------------------------------------------------------------')

    # Cierre de conexiones
    progress_bar.close()
    smtpObj.quit()
    mycursor.close()
    mydb.close()

if __name__ == "__main__":
    reverse_invoice_partial_ind_meli()
    reverse_invoice_partial_glob_meli()
    reverse_invoice_partial_ind_amz()
    reverse_invoice_partial_glo_amz()
    end_time = datetime.datetime.now()
    duration = end_time - today_date
    print(f'Duraciòn del script: {duration}')
    print('Listo')
    print('Este arroz ya se coció :)')